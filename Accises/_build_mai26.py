# -*- coding: utf-8 -*-
"""Reconstruit Mai-26 (déclaration principale, 11 lignes) dans le fichier AC4
et crée le fichier séparé AC4_Tea_Tree_SA_Mai26_partie1.xlsx (2ème déclaration)."""
import openpyxl, datetime
from openpyxl.styles import Font

T1, T2 = 0.681339, 1.192343
bold = Font(bold=True)

rows_mai = [
    ('Kirshner Fischer', 'RGK26-04142', datetime.datetime(2026, 5, 11), 40, 1),
    ('Kirshner Fischer', 'RGK26-04167', datetime.datetime(2026, 5, 12), 30, 1),
    ('Kirshner Fischer', 'RGK26-04188', datetime.datetime(2026, 5, 12), 40, 1),
    ('Kirshner Fischer', 'RGK26-04189', datetime.datetime(2026, 5, 12), 100, 1),
    ('Kirshner Fischer', 'RGK26-04267', datetime.datetime(2026, 5, 18), 170, 1),
    ('Kirshner Fischer', 'RGK26-04265', datetime.datetime(2026, 5, 19), 30, 1),
    ('Kirshner Fischer', 'RGK26-04266', datetime.datetime(2026, 5, 19), 49.7, 1),
    ('Kirshner Fischer', 'RGK26-04387', datetime.datetime(2026, 5, 20), 20, 1),
    ('Kirshner Fischer', 'RGK26-04484', datetime.datetime(2026, 5, 26), 25, 1),
    ('D&B', 'R1159008567771', datetime.datetime(2026, 5, 29), 80, 1),
    ('D&B', 'R1159008567771', datetime.datetime(2026, 5, 29), 30, 2),
]


def build_declaration_sheet(ws, titre, rows):
    ws['A1'] = titre; ws['G1'] = titre
    ws['L1'] = 'taux accises 1'; ws['M1'] = T1
    ws['A2'] = 'Entrées'; ws['G2'] = 'Sorties'
    headers = ['Fournisseurs', 'Facture n°', 'date', 'Kg de thé', 'Taux', 'Accises',
               'Fournisseurs', 'AC4 n°', 'date', 'Kg de thé', 'Accises']
    for c, v in zip('ABCDEFGHIJK', headers):
        ws[c + '3'] = v
    ws['L3'] = 'Taux accises 2'; ws['M3'] = T2
    for cell in ['A1', 'G1', 'A2', 'G2']:
        ws[cell].font = bold
    r = 4
    for four, fact, dt, kg, taux in rows:
        rate = T1 if taux == 1 else T2
        ws.cell(r, 1, four); ws.cell(r, 2, fact)
        c = ws.cell(r, 3, dt); c.number_format = 'DD/MM/YYYY'
        ws.cell(r, 4, kg); ws.cell(r, 5, taux); ws.cell(r, 6, kg * rate)
        ws.cell(r, 7, four)
        r += 1
    tot = r
    ws.cell(tot, 1, 'Total du mois').font = bold
    ws.cell(tot, 4, '=SUM(D4:D%d)' % (tot - 1)); ws.cell(tot, 6, '=SUM(F4:F%d)' % (tot - 1))
    ws.cell(tot, 7, 'Total du mois').font = bold
    ws.cell(tot, 10, '=SUM(J4:J%d)' % (tot - 1)); ws.cell(tot, 11, '=SUM(K4:K%d)' % (tot - 1))
    ws.cell(tot + 1, 3, 'Taux 1')
    ws.cell(tot + 1, 4, '=SUMIF(E4:E%d,1,D4:D%d)' % (tot - 1, tot - 1))
    ws.cell(tot + 1, 6, '=SUMIF(E4:E%d,1,F4:F%d)' % (tot - 1, tot - 1))
    ws.cell(tot + 2, 3, 'Taux 2')
    ws.cell(tot + 2, 4, '=SUMIF(E4:E%d,2,D4:D%d)' % (tot - 1, tot - 1))
    ws.cell(tot + 2, 6, '=SUMIF(E4:E%d,2,F4:F%d)' % (tot - 1, tot - 1))
    ws.cell(tot + 3, 4, '=D%d+D%d' % (tot + 1, tot + 2))
    ws.cell(tot + 3, 6, '=F%d+F%d' % (tot + 1, tot + 2))
    widths = {'A': 18, 'B': 16, 'C': 12, 'D': 10, 'E': 6, 'F': 12, 'G': 18,
              'H': 10, 'I': 12, 'J': 10, 'K': 10, 'L': 14, 'M': 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# 1. Fichier principal : restaurer Mai-26 sans les factures partie 1, retirer l'annexe
path = 'AC4_Tea_Tree_SA_Q126_updated v2.xlsx'
wb = openpyxl.load_workbook(path)
for s in ['Mai-26', 'Annexe Mai26 P1']:
    if s in wb.sheetnames:
        del wb[s]
ws = wb.create_sheet('Mai-26')
build_declaration_sheet(ws, 'Mai26', rows_mai)
wb.save(path)
print('Fichier principal restaure : Mai-26 = 11 lignes, annexe retiree')

# 2. Fichier séparé : déclaration partie 1
rows_p1 = [
    ('Kirshner Fischer', 'RGK26-03985', datetime.datetime(2026, 5, 6), 40, 1),
    ('Kirshner Fischer', 'RGK26-03998', datetime.datetime(2026, 5, 7), 10, 1),
]
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'Mai-26 partie 1'
build_declaration_sheet(ws2, 'Mai26 partie 1', rows_p1)

wa = wb2.create_sheet('Annexe détail')
wa['A1'] = 'ANNEXE DÉCLARATION ACCISES — Mai 2026 partie 1 (2ème déclaration)'
wa['A1'].font = bold
wa['A2'] = 'Tea Tree SA — Détail des factures RGK26-03985 et RGK26-03998 (Kirchner, Fischer & Co GmbH)'
hdr = ['Facture n°', 'Date facture', 'Pos', 'Article', 'Description fournisseur',
       'Réf. Teatower', 'Lot', 'Code tarifaire', 'Qté', 'Kg net thé',
       'Soumis accises', 'Taux', 'Accises €']
for i, h in enumerate(hdr, 1):
    c = wa.cell(4, i, h); c.font = bold

detail = [
    ('RGK26-03985', '06/05/2026', 1, '36100', 'Tea Ball Tong, stainless steel (Ø 4,5 cm)', 'A0274 Filtre pince', 'LOTK046785', '73239300900', '3 × 24 pce', 0, 'Non — accessoire', None),
    ('RGK26-03985', '06/05/2026', 2, '72046', 'Teapot "Globe" verre 1500 ml', 'A0378 Théière en verre Wide', 'LOTK042637', '70134200000', '15 pce', 0, 'Non — accessoire', None),
    ('RGK26-03985', '06/05/2026', 3, '26151', 'Apple/Vanilla — flavoured fruit tea blend', 'MP_1V0257 Pomme vanille', 'LOTK066688', '21069092859', '5 × 1 kg', 5, 'Oui', 1),
    ('RGK26-03985', '06/05/2026', 4, '18804', 'Fairy Dew — flavoured green tea', 'MP_1V0617 Liberté', 'LOTK070454', '09021000000', '10 × 1 kg', 10, 'Oui', 1),
    ('RGK26-03985', '06/05/2026', 5, '18728', 'My Lady — flavoured black tea blend', 'MP_1V0772 Black Mango', 'LOTK068738', '09023000000', '10 × 1 kg', 10, 'Oui', 1),
    ('RGK26-03985', '06/05/2026', 6, '18016', 'Turkish Honey/White Nougat — flavoured black tea', '18016', 'LOTK071198', '09023000000', '10 × 1 kg', 10, 'Oui', 1),
    ('RGK26-03985', '06/05/2026', 7, '25209', 'Mocca/Cream — flavoured Rooibos blend', 'MP_1V0874 Café liégeois', 'LOTK065642', '21069098690', '5 × 1 kg', 5, 'Oui', 1),
    ('RGK26-03985', '06/05/2026', 8, '1_VERS', 'Assurance', '', '', '', '1', 0, 'Non — frais', None),
    ('RGK26-03985', '06/05/2026', 9, '2_MAUT', 'Péage', '', '', '', '3', 0, 'Non — frais', None),
    ('RGK26-03998', '07/05/2026', 1, '388124', 'Fennel/Camomile — non-flavoured fruit/herbal tea blend', 'MP_1V0121 Lady Dodo', 'LOTK072410', '21069092859', '1 × 10 kg', 10, 'Oui', 1),
    ('RGK26-03998', '07/05/2026', 2, '2_MAUT', 'Péage', '', '', '', '1', 0, 'Non — frais', None),
]
r = 5
for fact, dt, pos, art, desc, ref, lot, tarif, qte, kg, soumis, taux in detail:
    wa.cell(r, 1, fact); wa.cell(r, 2, dt); wa.cell(r, 3, pos); wa.cell(r, 4, art)
    wa.cell(r, 5, desc); wa.cell(r, 6, ref); wa.cell(r, 7, lot); wa.cell(r, 8, tarif)
    wa.cell(r, 9, qte); wa.cell(r, 10, kg if kg else None); wa.cell(r, 11, soumis)
    if taux:
        wa.cell(r, 12, taux); wa.cell(r, 13, kg * T1)
    r += 1
r += 1
wa.cell(r, 1, 'Sous-total RGK26-03985').font = bold
wa.cell(r, 10, 40); wa.cell(r, 12, 1); wa.cell(r, 13, 40 * T1)
r += 1
wa.cell(r, 1, 'Sous-total RGK26-03998').font = bold
wa.cell(r, 10, 10); wa.cell(r, 12, 1); wa.cell(r, 13, 10 * T1)
r += 1
wa.cell(r, 1, 'TOTAL déclaration partie 1').font = bold
wa.cell(r, 10, 50).font = bold
wa.cell(r, 13, 50 * T1).font = bold
r += 2
wa.cell(r, 1, 'Taux accises 1 = %s €/kg (thé sans sucre ajouté) — Taux 2 = %s €/kg (avec sucre ajouté)' % (T1, T2))
widths = {'A': 22, 'B': 12, 'C': 5, 'D': 9, 'E': 46, 'F': 28, 'G': 12,
          'H': 14, 'I': 10, 'J': 11, 'K': 16, 'L': 6, 'M': 10}
for col, w in widths.items():
    wa.column_dimensions[col].width = w

wb2.save('AC4_Tea_Tree_SA_Mai26_partie1.xlsx')
print('Fichier separe cree : AC4_Tea_Tree_SA_Mai26_partie1.xlsx')
print('Partie 1 : 50 kg taux 1 = %.2f EUR' % (50 * T1))
print('Mai-26 (declaration principale) : 614,7 kg = %.2f EUR' % (584.7 * T1 + 30 * T2))

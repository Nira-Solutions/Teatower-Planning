"""Construit le feed Offer-only v6 pour les 205 SKU dont l'ASIN existe deja chez Amazon.

Source :
- _v6_sku_asin_map.json (SKU -> ASIN) extrait du processing-summary v5 erreur 8541
- data/GROCERY_Teatower_v5_ready.xlsx (prix + qty pour chaque SKU)

Sortie :
- data/Teatower_Offer_v6.tsv (Inventory Loader Amazon flat file, tab-delimited CP1252)

Format Amazon Inventory Loader minimal :
  sku  product-id  product-id-type  price  item-condition  quantity  add-delete
  product-id-type : 1=ASIN
  item-condition : 11=New
  add-delete : a=add
"""
import sys, io, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook

SKU_ASIN_JSON = './data/_v6_sku_asin_map.json'
V5_XLSX = './data/GROCERY_Teatower_v5_ready.xlsx'
OUT_TSV = './data/Teatower_Offer_v6.tsv'

with open(SKU_ASIN_JSON, 'r', encoding='utf-8') as f:
    sku_to_asin = json.load(f)
print(f'SKU charges depuis map : {len(sku_to_asin)}')

# Read v5 for price + qty per SKU
wb = load_workbook(V5_XLSX, data_only=True)
ws = wb['Modèle']
headers = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
hidx = {h: i for i, h in enumerate(headers) if h}

sku_col = hidx['item_sku']
# Price : list_price_with_tax (TTC)
price_col = hidx.get('list_price_with_tax')
# Qty : fulfillment_availability#1.quantity
qty_col = hidx.get('fulfillment_availability#1.quantity')

data = {}
for r in ws.iter_rows(min_row=4, values_only=True):
    sku = r[sku_col] if sku_col < len(r) else None
    if not sku: continue
    price = r[price_col] if price_col is not None and price_col < len(r) else None
    qty = r[qty_col] if qty_col is not None and qty_col < len(r) else None
    data[str(sku).strip()] = {'price': price, 'qty': qty}

print(f'SKU lus depuis v5 : {len(data)}')

# Build offer rows
HEADERS = ['sku', 'product-id', 'product-id-type', 'price', 'item-condition', 'quantity', 'add-delete']
rows = []
missing_price = []
missing_qty = []
for sku, asin in sku_to_asin.items():
    d = data.get(sku)
    if not d:
        print(f'  WARN : SKU {sku} absent de v5_ready.xlsx, skip')
        continue
    price = d['price']
    if not price:
        missing_price.append(sku)
        price = ''
    qty = d['qty']
    if not qty:
        missing_qty.append(sku)
        qty = '0'
    rows.append({
        'sku': sku,
        'product-id': asin,
        'product-id-type': '1',          # ASIN
        'price': str(price),
        'item-condition': '11',          # New
        'quantity': str(qty),
        'add-delete': 'a',
    })

print(f'\nLignes offer construites : {len(rows)}')
print(f'SKU sans prix : {len(missing_price)} -> {missing_price[:10]}')
print(f'SKU sans qty : {len(missing_qty)} -> {missing_qty[:10]}')

# Write TSV CP1252 - Inventory Loader format : headers on line 1, NO TemplateType preamble
# v6 erreur : avoir mis "TemplateType=Offer" en ligne 1 a fait qu'Amazon a lu cette ligne
# comme les en-tetes -> 90012 sku manquant + 7 warnings 90061 add-delete invalide.
with open(OUT_TSV, 'w', encoding='cp1252', errors='replace', newline='') as f:
    f.write('\t'.join(HEADERS) + '\r\n')
    for r in rows:
        f.write('\t'.join(r[h] for h in HEADERS) + '\r\n')

print(f'\nTSV sauvegarde : {OUT_TSV}')
print(f'Lignes : 1 header + {len(rows)} offers')

# Sample
print('\nSample 5 premieres offres :')
for r in rows[:5]:
    print(' ', r)

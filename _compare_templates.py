"""Compare old template (Teatower_Amazon_FBA_Ready) vs new GROCERY (1) template."""
import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OLD = './data/Teatower_Amazon_FBA_Ready.xlsx'
NEW = './data/GROCERY (1).xlsm'

# Old : headers technique = ligne 1
wb_o = openpyxl.load_workbook(OLD, read_only=True, data_only=True)
ws_o = wb_o['Amazon FBA - FR']
old_headers = [c.value or "" for c in next(ws_o.iter_rows(min_row=1, max_row=1))]
old_headers = [h for h in old_headers if h]

# New : Modèle sheet ligne 3
wb_n = openpyxl.load_workbook(NEW, read_only=True, data_only=True, keep_vba=False)
ws_n = wb_n['Modèle']
rows_n = list(ws_n.iter_rows(max_row=3, values_only=True))
new_meta = rows_n[0]
new_labels = rows_n[1]
new_headers = [h or "" for h in rows_n[2]]
new_headers = [h for h in new_headers if h]

print(f"OLD template : {len(old_headers)} colonnes")
print(f"NEW template : {len(new_headers)} colonnes")

old_set = set(old_headers)
new_set = set(new_headers)

print(f"\n=== Colonnes communes (mapping direct possible) : {len(old_set & new_set)} ===")
print(f"=== Colonnes SUPPRIMEES (dans l'ancien, plus dans le nouveau) : {len(old_set - new_set)} ===")
for h in sorted(old_set - new_set):
    print(f"  - {h}")

print(f"\n=== Colonnes NOUVELLES (dans le nouveau, pas dans l'ancien) : {len(new_set - old_set)} ===")
for h in sorted(new_set - old_set):
    print(f"  + {h}")

# Identifier les NEW columns critiques (obligatoires)
# On utilise la sheet "Définitions des données"
ws_def = wb_n['Définitions des données']
required = {}
for row in ws_def.iter_rows(min_row=3, values_only=True):
    field = row[1]  # Nom du champ technique
    req = row[6]    # Obligatoire ?
    if field and req:
        required[field] = str(req)

print(f"\n=== Champs OBLIGATOIRES (Requis) du nouveau template ===")
n_req = 0
for h in new_headers:
    r = required.get(h, "")
    if r and "equis" in r.lower():
        n_req += 1
        in_old = "OK (deja dans old)" if h in old_set else "** NOUVEAU OBLIGATOIRE **"
        print(f"  [{r}] {h}  {in_old}")
print(f"\nTotal champs Requis : {n_req}")

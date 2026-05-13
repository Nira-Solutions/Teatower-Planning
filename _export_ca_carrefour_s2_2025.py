"""Export Excel CA HTVA Carrefour Market/Hyper/Express - S2 2025."""
import xmlrpc.client, sys, io, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

URL = "https://tea-tree.odoo.com"
DB = "tsc-be-tea-tree-main-18515272"
USER = "nicolas.raes@teatower.com"
PWD = "Teatower123"

common = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/common")
uid = common.authenticate(DB, USER, PWD, {})
models = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/object")
def call(m, me, a, k=None): return models.execute_kw(DB, uid, PWD, m, me, a, k or {})

DATE_FROM = "2025-07-01"
DATE_TO = "2025-12-31"

# Partners Carrefour
all_carr = call("res.partner", "search_read",
    [[("name", "ilike", "carrefour")]],
    {"fields": ["id", "name"], "limit": 5000})
cm_extra = call("res.partner", "search_read",
    [["|", "|", ("name", "=ilike", "CM %"), ("name", "=ilike", "CE %"), ("name", "=ilike", "HC %")]],
    {"fields": ["id", "name"], "limit": 5000})
seen = {p["id"]: p for p in all_carr}
for p in cm_extra:
    seen.setdefault(p["id"], p)

def classify(name):
    n = name or ""
    if re.search(r"\bhyper(?:marche|marché|markt)?\s+carrefour\b|\bcarrefour\s+hyper\b|^\s*hc\s+", n, re.I):
        return "HYPER"
    if re.search(r"\bcarrefour\s+market\b|^\s*cm\s+", n, re.I):
        return "MARKET"
    if re.search(r"\bcarrefour\s+express\b|^\s*ce\s+", n, re.I):
        return "EXPRESS"
    return None

target = {pid: (p["name"], classify(p["name"])) for pid, p in seen.items() if classify(p["name"])}
ambiguous_ids = [pid for pid, p in seen.items() if not classify(p["name"])]

# SO scope strict
orders = call("sale.order", "search_read",
    [[("partner_id", "in", list(target.keys())),
      ("date_order", ">=", DATE_FROM),
      ("date_order", "<=", DATE_TO + " 23:59:59"),
      ("state", "in", ["sale", "done"])]],
    {"fields": ["name", "partner_id", "date_order", "amount_untaxed", "amount_total", "state", "user_id"],
     "order": "date_order asc"})

# SO ambigus
orders_amb = call("sale.order", "search_read",
    [[("partner_id", "in", ambiguous_ids),
      ("date_order", ">=", DATE_FROM),
      ("date_order", "<=", DATE_TO + " 23:59:59"),
      ("state", "in", ["sale", "done"])]],
    {"fields": ["name", "partner_id", "date_order", "amount_untaxed", "amount_total", "state", "user_id"],
     "order": "date_order asc"})

# === Build workbook ===
wb = Workbook()
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row):
    for c in ws[row]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# --- Sheet 1 : Synthese ---
ws1 = wb.active
ws1.title = "Synthese"
ws1["A1"] = "CA HTVA Carrefour Market / Hyper / Express"
ws1["A1"].font = Font(bold=True, size=14)
ws1["A2"] = f"Periode : {DATE_FROM}  ->  {DATE_TO}"
ws1["A2"].font = Font(italic=True, color="595959")
ws1["A3"] = "Source : Odoo Teatower - sale.order (state in sale/done)"
ws1["A3"].font = Font(italic=True, color="595959")

ws1.append([])
ws1.append(["Segment", "Nb commandes", "CA HTVA (EUR)", "CA TTC (EUR)"])
style_header(ws1, ws1.max_row)

ht_by_type = {"MARKET": 0, "HYPER": 0, "EXPRESS": 0}
ttc_by_type = {"MARKET": 0, "HYPER": 0, "EXPRESS": 0}
n_by_type = {"MARKET": 0, "HYPER": 0, "EXPRESS": 0}
for o in orders:
    t = target[o["partner_id"][0]][1]
    ht_by_type[t] += o["amount_untaxed"]
    ttc_by_type[t] += o["amount_total"]
    n_by_type[t] += 1

for seg in ["MARKET", "HYPER", "EXPRESS"]:
    ws1.append([f"Carrefour {seg.title()}", n_by_type[seg], ht_by_type[seg], ttc_by_type[seg]])
ws1.append(["TOTAL", sum(n_by_type.values()), sum(ht_by_type.values()), sum(ttc_by_type.values())])
total_row = ws1.max_row
for c in ws1[total_row]:
    c.font = Font(bold=True)
    c.fill = TOTAL_FILL
    c.border = BORDER

# Format euros
for r in range(6, total_row + 1):
    for col in ("C", "D"):
        ws1[f"{col}{r}"].number_format = '#,##0.00 "EUR"'
        ws1[f"{col}{r}"].border = BORDER
    ws1[f"A{r}"].border = BORDER
    ws1[f"B{r}"].border = BORDER

ws1.append([])
ws1.append(["NOTES"])
ws1[ws1.max_row][0].font = Font(bold=True)
ws1.append(["Partners 'Carrefour' exclus du scope :", "voir onglet 'Cas ambigus'"])
ws1.append(["DEDROP - Brasserie le Carrefour", "Restaurant independant, rien a voir avec l'enseigne -> exclu"])
ws1.append(["Carrefour Belgium - Corporate Village", "Probable cafeteria siege - a statuer"])
ws1.append(["UNIC S.A. Carrefour Florenville", "Probable Carrefour Market - a statuer"])

autosize(ws1, [45, 16, 18, 18])

# --- Sheet 2 : Par magasin ---
ws2 = wb.create_sheet("Par magasin")
ws2.append(["Segment", "Magasin (partner Odoo)", "ID partner", "Nb cmd", "CA HTVA (EUR)", "CA TTC (EUR)"])
style_header(ws2, 1)

by_partner = {}
for o in orders:
    pid = o["partner_id"][0]
    by_partner.setdefault(pid, {"n": 0, "ht": 0, "ttc": 0})
    by_partner[pid]["n"] += 1
    by_partner[pid]["ht"] += o["amount_untaxed"]
    by_partner[pid]["ttc"] += o["amount_total"]

for pid, d in sorted(by_partner.items(), key=lambda x: -x[1]["ht"]):
    name, t = target[pid]
    ws2.append([t, name, pid, d["n"], d["ht"], d["ttc"]])

# Total row
ws2.append(["TOTAL", "", "", sum(d["n"] for d in by_partner.values()),
            sum(d["ht"] for d in by_partner.values()),
            sum(d["ttc"] for d in by_partner.values())])
total_row2 = ws2.max_row
for c in ws2[total_row2]:
    c.font = Font(bold=True)
    c.fill = TOTAL_FILL
for r in range(2, total_row2 + 1):
    ws2[f"E{r}"].number_format = '#,##0.00 "EUR"'
    ws2[f"F{r}"].number_format = '#,##0.00 "EUR"'
    for col in "ABCDEF":
        ws2[f"{col}{r}"].border = BORDER

autosize(ws2, [10, 55, 10, 8, 16, 16])

# --- Sheet 3 : Detail SO ---
ws3 = wb.create_sheet("Detail SO")
ws3.append(["SO", "Date", "Segment", "Magasin", "ID partner", "Vendeur", "Etat", "Montant HTVA", "Montant TTC"])
style_header(ws3, 1)

for o in orders:
    pid = o["partner_id"][0]
    name, t = target[pid]
    vendeur = o["user_id"][1] if o.get("user_id") else ""
    ws3.append([o["name"], o["date_order"][:10], t, name, pid, vendeur, o["state"],
                o["amount_untaxed"], o["amount_total"]])

total_row3 = ws3.max_row + 1
ws3.cell(row=total_row3, column=1, value="TOTAL").font = Font(bold=True)
ws3.cell(row=total_row3, column=8, value=sum(o["amount_untaxed"] for o in orders)).font = Font(bold=True)
ws3.cell(row=total_row3, column=9, value=sum(o["amount_total"] for o in orders)).font = Font(bold=True)
for c in ws3[total_row3]:
    c.fill = TOTAL_FILL

for r in range(2, total_row3 + 1):
    ws3[f"H{r}"].number_format = '#,##0.00 "EUR"'
    ws3[f"I{r}"].number_format = '#,##0.00 "EUR"'
    for col in "ABCDEFGHI":
        ws3[f"{col}{r}"].border = BORDER

autosize(ws3, [10, 12, 10, 50, 10, 22, 8, 14, 14])

# --- Sheet 4 : Cas ambigus ---
ws4 = wb.create_sheet("Cas ambigus")
ws4["A1"] = "Partners 'Carrefour' non classes Market/Hyper/Express - SO S2 2025"
ws4["A1"].font = Font(bold=True, size=12)
ws4["A2"] = "Confirmer avec Nicolas pour eventuelle reinclusion dans le scope"
ws4["A2"].font = Font(italic=True, color="C00000")

ws4.append([])
ws4.append(["SO", "Date", "Partner", "ID partner", "Etat", "HTVA", "TTC"])
style_header(ws4, ws4.max_row)

# Group ambigus
amb_by_partner = {}
for o in orders_amb:
    pid = o["partner_id"][0]
    amb_by_partner.setdefault(pid, {"name": o["partner_id"][1], "n": 0, "ht": 0, "ttc": 0, "orders": []})
    amb_by_partner[pid]["n"] += 1
    amb_by_partner[pid]["ht"] += o["amount_untaxed"]
    amb_by_partner[pid]["ttc"] += o["amount_total"]
    amb_by_partner[pid]["orders"].append(o)

# Synthese ambigus
ws4.append([])
row_synth = ws4.max_row + 1
ws4.cell(row=row_synth, column=1, value="SYNTHESE PAR PARTNER AMBIGU").font = Font(bold=True)
ws4.append(["Partner", "ID partner", "Nb SO", "CA HTVA", "CA TTC", "Decision suggeree", ""])
style_header(ws4, ws4.max_row)

suggestions = {
    6596: "EXCLURE - cafet siege Carrefour Belgium",
    2903: "EXCLURE - brasserie independante",
    123089: "EXCLURE - entite centrale",
    3272: "INCLURE en MARKET (Florenville)",
    6098: "EXCLURE - partner generique",
    9423: "EXCLURE - email mal saisi",
}
for pid, d in sorted(amb_by_partner.items(), key=lambda x: -x[1]["ht"]):
    ws4.append([d["name"], pid, d["n"], d["ht"], d["ttc"], suggestions.get(pid, "A statuer")])
total_amb_row = ws4.max_row + 1
ws4.cell(row=total_amb_row, column=1, value="TOTAL ambigus").font = Font(bold=True)
ws4.cell(row=total_amb_row, column=3, value=sum(d["n"] for d in amb_by_partner.values())).font = Font(bold=True)
ws4.cell(row=total_amb_row, column=4, value=sum(d["ht"] for d in amb_by_partner.values())).font = Font(bold=True)
ws4.cell(row=total_amb_row, column=5, value=sum(d["ttc"] for d in amb_by_partner.values())).font = Font(bold=True)

# Format
for r in range(row_synth, total_amb_row + 1):
    for col in "DE":
        ws4[f"{col}{r}"].number_format = '#,##0.00 "EUR"'

autosize(ws4, [45, 12, 8, 14, 14, 40, 5])

# --- Save ---
out = f"./CA_Carrefour_S2-2025_{DATE_FROM}_{DATE_TO}.xlsx"
wb.save(out)
print(f"Excel sauvegarde : {out}")
print(f"Onglets : Synthese, Par magasin, Detail SO, Cas ambigus")
print(f"Total scope strict : {sum(ht_by_type.values()):,.2f} EUR HT ({sum(n_by_type.values())} SO)")
print(f"Total cas ambigus  : {sum(d['ht'] for d in amb_by_partner.values()):,.2f} EUR HT ({sum(d['n'] for d in amb_by_partner.values())} SO)")

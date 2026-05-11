"""
Export Excel pour Vanessa (support des ventes) — Campagne juin 2026.

Génère un xlsx avec :
- 1 onglet "Synthèse" : récap des 3 segments
- 1 onglet par segment (Horeca actif, Horeca dormant, Revendeur dormant)
- Colonnes opérationnelles : partner_id, code promo, nom, contact, email, tél/mobile,
  adresse, dernière commande, CA 12 mois, statut, notes

Source : tags MC-202606-* déjà appliqués dans Odoo.
"""
import xmlrpc.client, sys, io, hashlib, datetime as dt
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl manquant — pip install openpyxl")
    sys.exit(1)

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

URL = "https://tea-tree.odoo.com"
DB = "tsc-be-tea-tree-main-18515272"
USER = "nicolas.raes@teatower.com"
PWD = "Teatower123"

TAGS = {
    "MC-202606-HORECA-ACTIF":      ("HORECA ACTIF",      "Cafés/restos qui commandent activement"),
    "MC-202606-HORECA-DORMANT":    ("HORECA DORMANT",    "Sans commande > 6 mois — relance Jérôme"),
    "MC-202606-REVENDEUR-DORMANT": ("REVENDEUR DORMANT", "Sans commande > 6 mois — offre -10%"),
}

OUT_XLSX = Path(__file__).parent / "Liste_Vanessa_Juin_2026.xlsx"

# --- ODOO CONNECTION ---
common = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/common")
uid = common.authenticate(DB, USER, PWD, {})
models = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/object")

def call(model, method, args, kwargs=None):
    return models.execute_kw(DB, uid, PWD, model, method, args, kwargs or {})

def gen_promo_code(partner_id):
    h = hashlib.md5(f"teatower-{partner_id}-202606".encode()).hexdigest()
    return f"TT-J26-{h[:6].upper()}"

# --- 1. Find the 3 tags ---
tag_ids = {}
for tag_name in TAGS:
    res = call("res.partner.category", "search", [[("name", "=", tag_name)]], {"limit": 1})
    if not res:
        print(f"ATTENTION : tag {tag_name} introuvable dans Odoo. Lancer d'abord _tag_campagne_juin_2026.py.")
        sys.exit(1)
    tag_ids[tag_name] = res[0]
print(f"Tags trouves: {tag_ids}")

# --- 2. Fetch partners for each tag ---
all_partners = {}  # partner_id -> data dict
for tag_name, tag_id in tag_ids.items():
    partners = call("res.partner", "search_read",
        [[("category_id", "in", [tag_id])]],
        {"fields": ["id", "name", "email", "phone", "mobile",
                    "street", "street2", "zip", "city", "country_id",
                    "vat", "ref", "comment"],
         "limit": 5000})
    for p in partners:
        p["_segment"] = TAGS[tag_name][0]
        p["_segment_desc"] = TAGS[tag_name][1]
        p["_promo_code"] = gen_promo_code(p["id"])
        all_partners[p["id"]] = p
    print(f"  {tag_name}: {len(partners)} partners")

partner_ids = list(all_partners.keys())
print(f"\nTotal partners tagges: {len(partner_ids)}")

# --- 3. Fetch last order + CA 12 months per partner ---
print("Calcul derniere commande + CA 12 mois ...")
date_from = (dt.date.today() - dt.timedelta(days=365)).isoformat()

orders_all = call("sale.order", "search_read",
    [[("partner_id", "in", partner_ids),
      ("state", "in", ["sale", "done"])]],
    {"fields": ["partner_id", "date_order", "amount_untaxed", "amount_total"],
     "order": "date_order desc"})

last_order = {}
ca_12m = {}
nb_orders_12m = {}
for o in orders_all:
    pid = o["partner_id"][0]
    if pid not in last_order:
        last_order[pid] = o["date_order"][:10]
    if o["date_order"][:10] >= date_from:
        ca_12m[pid] = ca_12m.get(pid, 0) + o["amount_untaxed"]
        nb_orders_12m[pid] = nb_orders_12m.get(pid, 0) + 1

# --- 4. Fetch primary contact (child_ids) per partner ---
print("Recuperation contacts principaux ...")
contacts_per_partner = {}
contacts = call("res.partner", "search_read",
    [[("parent_id", "in", partner_ids), ("type", "in", ["contact", False])]],
    {"fields": ["id", "name", "function", "email", "phone", "mobile", "parent_id"],
     "limit": 5000})
for c in contacts:
    pid = c["parent_id"][0]
    if pid not in contacts_per_partner:
        contacts_per_partner[pid] = c

# --- 5. Build Excel ---
wb = openpyxl.Workbook()
header_font = Font(bold=True, color="FFFFFF", size=10)
header_fill = PatternFill("solid", fgColor="0C4A30")
header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
even_fill = PatternFill("solid", fgColor="F8F5EF")
border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

COLS = [
    ("Code promo",        20, "_promo_code"),
    ("Segment",           20, "_segment"),
    ("Société",           40, "name"),
    ("Contact",           25, "_contact_name"),
    ("Fonction",          20, "_contact_function"),
    ("Email société",     30, "email"),
    ("Email contact",     30, "_contact_email"),
    ("Tél fixe",          16, "_phone_combined"),
    ("Mobile",            16, "_mobile_combined"),
    ("Adresse",           35, "_address"),
    ("Ville",             20, "city"),
    ("Dernière commande", 14, "_last_order"),
    ("CA 12M (€)",        12, "_ca_12m"),
    ("Nb cmds 12M",       10, "_nb_orders_12m"),
    ("ID Odoo",           10, "id"),
    ("Notes",             30, "_notes"),
]

def enrich(p):
    """Ajoute les champs calculés."""
    c = contacts_per_partner.get(p["id"], {})
    p["_contact_name"]     = c.get("name") or ""
    p["_contact_function"] = c.get("function") or ""
    p["_contact_email"]    = c.get("email") or ""
    # Combined phone : société + contact
    phones = [x for x in [p.get("phone"), c.get("phone")] if x]
    mobiles = [x for x in [p.get("mobile"), c.get("mobile")] if x]
    p["_phone_combined"]  = " / ".join(set(phones))
    p["_mobile_combined"] = " / ".join(set(mobiles))
    # Address
    addr_parts = [p.get("street") or "", p.get("street2") or "", p.get("zip") or ""]
    p["_address"] = " ".join([x for x in addr_parts if x]).strip()
    p["_last_order"]    = last_order.get(p["id"], "(jamais)")
    p["_ca_12m"]        = round(ca_12m.get(p["id"], 0), 2)
    p["_nb_orders_12m"] = nb_orders_12m.get(p["id"], 0)
    p["_notes"]         = ""  # vide, à remplir par Vanessa
    return p

def write_sheet(ws, partners_list, title):
    ws.title = title
    # Headers
    for col_idx, (label, width, _) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28
    # Data
    for row_idx, p in enumerate(partners_list, start=2):
        for col_idx, (label, _, field) in enumerate(COLS, start=1):
            val = p.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = even_fill
    # Freeze header
    ws.freeze_panes = "A2"
    # Autofilter
    ws.auto_filter.ref = ws.dimensions

# Enrich all partners
for pid, p in all_partners.items():
    enrich(p)

# Synthèse sheet first
ws_synthese = wb.active
ws_synthese.title = "Synthèse"
ws_synthese["A1"] = "CAMPAGNE JUIN 2026 — Suivi Vanessa"
ws_synthese["A1"].font = Font(bold=True, size=14, color="0C4A30")
ws_synthese.merge_cells("A1:E1")

ws_synthese["A3"] = "Segment"
ws_synthese["B3"] = "Volume"
ws_synthese["C3"] = "Avec email"
ws_synthese["D3"] = "CA cible"
ws_synthese["E3"] = "Offre"
for col in "ABCDE":
    c = ws_synthese[f"{col}3"]
    c.font = header_font
    c.fill = header_fill
    c.alignment = header_align
    c.border = border

# Counts per segment
seg_partners = {label: [] for (label, desc) in TAGS.values()}
for p in all_partners.values():
    seg_partners[p["_segment"]].append(p)

synth_rows = [
    ("HORECA ACTIF",      seg_partners["HORECA ACTIF"],      "+5 K€",  "3+1 sur boîtes HC25"),
    ("HORECA DORMANT",    seg_partners["HORECA DORMANT"],    "+1 K€",  "3+1 HC25 + relance Jérôme"),
    ("REVENDEUR DORMANT", seg_partners["REVENDEUR DORMANT"], "+8 K€",  "-10% sur réassort libre ≥ 300€"),
]
total_vol = 0
total_email = 0
for row_idx, (seg, plist, ca_cible, offre) in enumerate(synth_rows, start=4):
    vol = len(plist)
    nb_email = sum(1 for p in plist if (p.get("email") or "").strip() and "@" in (p.get("email") or ""))
    total_vol += vol
    total_email += nb_email
    ws_synthese[f"A{row_idx}"] = seg
    ws_synthese[f"B{row_idx}"] = vol
    ws_synthese[f"C{row_idx}"] = nb_email
    ws_synthese[f"D{row_idx}"] = ca_cible
    ws_synthese[f"E{row_idx}"] = offre
    for col in "ABCDE":
        ws_synthese[f"{col}{row_idx}"].border = border
        ws_synthese[f"{col}{row_idx}"].alignment = Alignment(vertical="top", wrap_text=True)

ws_synthese[f"A{4+len(synth_rows)}"] = "TOTAL"
ws_synthese[f"B{4+len(synth_rows)}"] = total_vol
ws_synthese[f"C{4+len(synth_rows)}"] = total_email
ws_synthese[f"D{4+len(synth_rows)}"] = "+14 K€"
for col in "ABCD":
    cell = ws_synthese[f"{col}{4+len(synth_rows)}"]
    cell.font = Font(bold=True)
    cell.border = border

# Notes / mode d'emploi
notes_start = 4 + len(synth_rows) + 3
ws_synthese[f"A{notes_start}"] = "MODE D'EMPLOI"
ws_synthese[f"A{notes_start}"].font = Font(bold=True, color="0C4A30", size=12)
notes = [
    "",
    "Cette campagne est envoyée par Stephan via Mailchimp le 9 juin 2026 à 9h00.",
    "",
    "TON RÔLE Vanessa :",
    "1. Quand un client commande en mentionnant son code TT-J26-XXXXXX → saisir le code dans le champ 'Référence client' du devis Odoo.",
    "2. Si client appelle pour bénéficier de l'offre sans avoir reçu l'email : retrouver son code dans ce fichier (Ctrl+F), confirmer l'offre.",
    "3. Si un email bounce/n'est plus valide : mettre à jour la fiche partner dans Odoo + noter en colonne 'Notes'.",
    "4. Si un client refuse la prochaine campagne : noter en 'Notes' (et signaler à Nicolas pour exclusion future).",
    "",
    "TES INTERLOCUTEURS :",
    "• Stephan = envoi Mailchimp",
    "• Jérôme = relance commerciale top 20 Revendeur + top 5 Horeca dormants à J+10 (19 juin)",
    "• Nicolas = reporting J+30 (9 juillet)",
    "",
    "FORMULE PROMO :",
    "• Horeca (actif + dormant) : 3 boîtes HC25 achetées = 4ᵉ offerte (min 6 boîtes, port offert 200€)",
    "• Revendeur dormant : -10% sur toute commande ≥ 300€ (1 commande max par client, port offert 250€)",
]
for i, txt in enumerate(notes):
    cell = ws_synthese.cell(row=notes_start + 1 + i, column=1, value=txt)
    if txt.startswith(("TON RÔLE", "TES INTERLOCUTEURS", "FORMULE PROMO")):
        cell.font = Font(bold=True, color="6B4423")
    ws_synthese.merge_cells(start_row=notes_start + 1 + i, start_column=1,
                             end_row=notes_start + 1 + i, end_column=8)

ws_synthese.column_dimensions["A"].width = 20
ws_synthese.column_dimensions["B"].width = 12
ws_synthese.column_dimensions["C"].width = 12
ws_synthese.column_dimensions["D"].width = 12
ws_synthese.column_dimensions["E"].width = 50

# Sheet per segment
order = ["HORECA ACTIF", "HORECA DORMANT", "REVENDEUR DORMANT"]
sheet_names = {
    "HORECA ACTIF":      "Horeca actif (3+1)",
    "HORECA DORMANT":    "Horeca dormant (3+1)",
    "REVENDEUR DORMANT": "Revendeur dormant (-10%)",
}
for seg in order:
    plist = sorted(seg_partners[seg], key=lambda p: (-p["_ca_12m"], p["name"] or ""))
    ws = wb.create_sheet(title=sheet_names[seg])
    write_sheet(ws, plist, sheet_names[seg])

# Onglet "Toutes" agrégé
plist_all = sorted(all_partners.values(),
                   key=lambda p: (p["_segment"], -p["_ca_12m"], p["name"] or ""))
ws_all = wb.create_sheet(title="Toutes campagnes")
write_sheet(ws_all, plist_all, "Toutes campagnes")

wb.save(OUT_XLSX)
print(f"\nExcel ecrit : {OUT_XLSX}")
print(f"  Taille : {OUT_XLSX.stat().st_size / 1024:.1f} Ko")
print(f"  Onglets : Synthese + 3 segments + Toutes")
print(f"  Total clients : {len(all_partners)}")

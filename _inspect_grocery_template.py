"""Inspect GROCERY (1).xlsm template structure - identify header rows and columns."""
import openpyxl
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

src = './data/GROCERY (1).xlsm'
wb = openpyxl.load_workbook(src, read_only=True, data_only=True, keep_vba=False)

print(f"Sheets: {wb.sheetnames}\n")
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"--- Sheet: {sn} | rows={ws.max_row} cols={ws.max_column} ---")
    # First 6 rows preview (first 8 cols)
    for r_idx, row in enumerate(ws.iter_rows(max_row=6, max_col=8, values_only=True), 1):
        vals = [str(v)[:30] if v is not None else "" for v in row]
        print(f"  L{r_idx}: {vals}")
    print()

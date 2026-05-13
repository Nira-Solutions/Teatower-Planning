"""Remap 335 products from old template to new GROCERY (1).xlsm schema 2026.
Output:
  - data/GROCERY_Teatower_v4_ready.xlsx (xlsx editable)
  - data/GROCERY_Teatower_v4_ready.txt  (Latin-1 tab-delimited for Amazon upload)
"""
import openpyxl
import unicodedata
import csv
import sys, io
import shutil
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OLD = './data/Teatower_Amazon_FBA_Ready.xlsx'
NEW_TEMPLATE = './data/GROCERY (1).xlsm'
OUT_XLSX = './data/GROCERY_Teatower_v4_ready.xlsx'
OUT_TXT = './data/GROCERY_Teatower_v4_ready.txt'

TVA_FR_THE = 1.055  # taux reduit 5,5% the/epicerie

# Marketplaces (5)
MARKETPLACES = [
    "A13V1IB3VIYZZH",  # FR
    "A1PA6795UKMFR9",  # DE
    "A1RKKUPIHCS9HS",  # NL
    "A1805IZSGTT6HS",  # IT
    "AMEN7PMS3EDWL",   # BE/SE
]

# Defaults pour nouveaux champs obligatoires (the - pas de batterie, pas dangereux)
DEFAULTS = {
    "batteries_required": "No",
    "are_batteries_included": "No",
    "battery_weight": "0",
    "battery_weight_unit_of_measure": "",
    "battery_cell_composition": "",
    "lithium_battery_weight": "0",
    "lithium_battery_weight_unit_of_measure": "",
    "lithium_battery_energy_content": "0",
    "lithium_battery_energy_content_unit_of_measure": "",
    "lithium_battery_packaging": "",
    "number_of_lithium_ion_cells": "0",
    "number_of_lithium_metal_cells": "0",
    "hazmat_united_nations_regulatory_id": "",
    "contains_liquid_contents": "No",
    "is_heat_sensitive": "No",
    "contains_food_or_beverage": "Yes",
    "temperature_rating": "Ambiante : Température ambiante",
    "each_unit_count": "1",
    "item_volume": "",
    "item_volume_unit_of_measure": "",
    "safety_data_sheet_url": "",
    "fulfillment_availability#1.fulfillment_channel_code": "DEFAULT",  # FBM
    "fulfillment_availability#1.is_inventory_available": "true",
    "fulfillment_availability#1.lead_time_to_ship_max_days": "3",
}

# Nettoyage caracteres Unicode -> ASCII-safe
SMART_QUOTE_MAP = str.maketrans({
    "‘": "'",  "’": "'",
    "“": '"',  "”": '"',
    "–": "-",  "—": "-",
    "…": "...",
    " ": " ",
    " ": " ",
    " ": " ",
    "​": "",
    "­": "",
    "﻿": "",
})

def is_emoji_or_symbol(ch):
    cp = ord(ch)
    if 0x2300 <= cp <= 0x27BF: return True       # symbols, dingbats, misc tech
    if 0x1F000 <= cp <= 0x1FFFF: return True     # all emoji blocks
    if 0xFE00 <= cp <= 0xFE0F: return True       # variation selectors
    if 0x200C <= cp <= 0x200F: return True       # ZWJ etc.
    return False

def clean_text(v):
    if v is None:
        return ""
    s = str(v)
    s = s.translate(SMART_QUOTE_MAP)
    s = "".join("" if is_emoji_or_symbol(c) else c for c in s)
    s = unicodedata.normalize("NFC", s)
    s = s.replace("\t", " ").replace("\r", " ").replace("\n", " ")
    while "  " in s:
        s = s.replace("  ", " ")
    return s.strip()

def to_cp1252_safe(s):
    """Force Windows-1252 compatibility (Amazon expects CP1252, not pure Latin-1)."""
    try:
        s.encode("cp1252")
        return s
    except UnicodeEncodeError:
        out = []
        for ch in s:
            try:
                ch.encode("cp1252")
                out.append(ch)
            except UnicodeEncodeError:
                nf = unicodedata.normalize("NFKD", ch)
                ascii_part = nf.encode("ascii", "ignore").decode("ascii")
                out.append(ascii_part if ascii_part else "")
        return "".join(out)

# === Load NEW template ===
shutil.copy(NEW_TEMPLATE, OUT_XLSX)  # base copy with all sheets
wb_out = load_workbook(OUT_XLSX, keep_vba=False)
# Drop unused sheets except Modele to keep file lean? No, keep all for user reference.
ws_mod = wb_out['Modèle']

# Read row 3 (technical headers)
new_headers = []
for c in next(ws_mod.iter_rows(min_row=3, max_row=3, values_only=True)):
    new_headers.append(c if c else "")
# Trim trailing empties
while new_headers and not new_headers[-1]:
    new_headers.pop()

print(f"New template : {len(new_headers)} colonnes")

# Build header -> col index map (1-based for openpyxl)
hidx = {h: i + 1 for i, h in enumerate(new_headers)}

# === Load OLD data ===
wb_old = openpyxl.load_workbook(OLD, read_only=True, data_only=True)
ws_old = wb_old['Amazon FBA - FR']
old_rows = list(ws_old.iter_rows(values_only=True))
old_headers = [h or "" for h in old_rows[0]]
# Detect label-row (row 2 has FR labels - skip)
data_start = 2  # row index 0 = tech headers, 1 = FR labels, 2+ = data

# Build old col map
old_idx = {h: i for i, h in enumerate(old_headers) if h}

# Mapping renames old -> new (cases ou le nom a change)
RENAME = {
    "quantity": "fulfillment_availability#1.quantity",
    # Si standard_price etait dans old, on l'utilise pour list_price_with_tax (calc).
}

# Find old price column
PRICE_OLD_COL = "standard_price" if "standard_price" in old_idx else None

n_skus = 0
skipped_empty = 0
sku_tea_conflict = []  # I0735 I0723 I0628 etc.
sku_weight_fixed = []  # HH003 HH004
unmappable = set()

# Iterate over old products starting row 3 (index 2)
out_row = 4  # template L1, L2, L3 reserves ; data L4+
for row_values in old_rows[data_start:]:
    if not row_values:
        skipped_empty += 1
        continue
    sku_v = row_values[old_idx.get("item_sku", 1)] if old_idx.get("item_sku") is not None else None
    if not sku_v:
        skipped_empty += 1
        continue
    n_skus += 1
    sku = clean_text(sku_v)

    # 1) Direct mapping for common cols
    out_cells = [""] * len(new_headers)
    for new_h, new_pos in hidx.items():
        if new_h in old_idx:
            v = row_values[old_idx[new_h]]
            out_cells[new_pos - 1] = clean_text(v)

    # 2) Renames
    for old_h, new_h in RENAME.items():
        if old_h in old_idx and new_h in hidx:
            v = row_values[old_idx[old_h]]
            out_cells[hidx[new_h] - 1] = clean_text(v)

    # 3) Defaults pour nouveaux champs obligatoires
    for k, default_v in DEFAULTS.items():
        if k in hidx:
            cur = out_cells[hidx[k] - 1]
            if not cur:
                out_cells[hidx[k] - 1] = default_v

    # 4) Calcul prix TTC depuis standard_price
    price_ht = None
    if PRICE_OLD_COL:
        raw_price = row_values[old_idx[PRICE_OLD_COL]]
        try:
            price_ht = float(raw_price)
        except (TypeError, ValueError):
            price_ht = None

    if price_ht is not None:
        price_ttc = round(price_ht * TVA_FR_THE, 2)
        if "list_price_with_tax" in hidx:
            out_cells[hidx["list_price_with_tax"] - 1] = f"{price_ttc:.2f}"
        # Apply to 5 marketplaces our_price
        for mp in MARKETPLACES:
            k = f"purchasable_offer[marketplace_id={mp}]#1.our_price#1.schedule#1.value_with_tax"
            if k in hidx:
                out_cells[hidx[k] - 1] = f"{price_ttc:.2f}"

    # 5) item_weight : limiter a 2 decimales
    if "item_weight" in hidx:
        v = out_cells[hidx["item_weight"] - 1]
        if v:
            try:
                f = float(v)
                v2 = f"{round(f, 2)}"
                out_cells[hidx["item_weight"] - 1] = v2
                if abs(f - round(f, 2)) > 1e-9:
                    sku_weight_fixed.append((sku, v, v2))
            except ValueError:
                pass

    # 6) feed_product_type forcer "grocery" (template GROCERY)
    if "feed_product_type" in hidx:
        cur = out_cells[hidx["feed_product_type"] - 1]
        if cur.lower() not in ("grocery", ""):
            # Note conflit potentiel TEA si SKU deja en ASIN TEA
            pass
        out_cells[hidx["feed_product_type"] - 1] = "grocery"

    # 7) Detect SKU I0xxx en conflit TEA (vu dans summary precedent)
    if sku in ("I0735", "I0723", "I0628"):
        sku_tea_conflict.append(sku)

    # Write row
    for col_i, val in enumerate(out_cells, 1):
        ws_mod.cell(row=out_row, column=col_i, value=val)
    out_row += 1

print(f"\nSKU exportes : {n_skus}")
print(f"Lignes vides ignorees : {skipped_empty}")
print(f"SKU avec conflit TEA (vont planter, a migrer manuellement) : {sku_tea_conflict}")
print(f"SKU avec item_weight arrondi (2 decimales) : {len(sku_weight_fixed)}")
for s, old, new in sku_weight_fixed[:10]:
    print(f"  {s}: {old} -> {new}")

# Save xlsx
wb_out.save(OUT_XLSX)
print(f"\nXLSX sauvegarde : {OUT_XLSX}")

# === Export .txt tab-delimited CP1252 (Windows-1252) ===
wb2 = load_workbook(OUT_XLSX, data_only=True)
ws2 = wb2['Modèle']

n_lines = 0
n_replaced = 0
with open(OUT_TXT, "w", encoding="cp1252", errors="replace", newline="") as f:
    for row in ws2.iter_rows(values_only=True):
        cells = []
        for v in row:
            if v is None:
                cells.append("")
            else:
                if isinstance(v, (int, float)):
                    s = str(v)
                else:
                    s = clean_text(v)
                s_safe = to_cp1252_safe(s)
                if s_safe != s:
                    n_replaced += 1
                s_safe = s_safe.replace("\t", " ").replace("\r", " ").replace("\n", " ")
                cells.append(s_safe)
        while cells and cells[-1] == "":
            cells.pop()
        f.write("\t".join(cells))
        f.write("\r\n")
        n_lines += 1

print(f"TXT sauvegarde : {OUT_TXT}")
print(f"Lignes ecrites : {n_lines}")
print(f"Caracteres non Latin-1 remplaces : {n_replaced}")

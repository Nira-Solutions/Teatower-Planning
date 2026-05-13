"""Remap Teatower SKU vers schema GROCERY 2026 (v5).

Differences vs v4 :
- DEFAULTS : lithium cells/weight = vides (pas "0") -> fix 702 erreurs 90114
- Ajoute item_length/width/height_unit_of_measure = centimeters -> fix 1404 warnings 99005
- Tronque flavor_name a 50 chars (smart : partie apres dernier '-' si dispo) -> fix 125 erreurs 90117
- Arrondit package_weight a 2 decimales -> fix 2 erreurs 90115
- Retire variation_theme / parent_sku / parent_child / relationship_type / size_name -> passe en SKU simples (fix 410 erreurs 99003 + 202 erreurs 8007)
- Skip les 10 SKU bloques (4 TEA conflit + 6 brand registry) :
  I0735, I0723, I0628, I0205, C0184, I0669

Output:
  - data/GROCERY_Teatower_v5_ready.xlsx (xlsx editable)
  - data/GROCERY_Teatower_v5_ready.txt  (CP1252 tab-delimited pour Seller Central)
"""
import openpyxl
import unicodedata
import sys, io
import shutil
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

OLD = './data/Teatower_Amazon_FBA_Ready.xlsx'
NEW_TEMPLATE = './data/GROCERY (1).xlsm'
OUT_XLSX = './data/GROCERY_Teatower_v5_ready.xlsx'
OUT_TXT = './data/GROCERY_Teatower_v5_ready.txt'

TVA_FR_THE = 1.055

MARKETPLACES = [
    "A13V1IB3VIYZZH",  # FR
    "A1PA6795UKMFR9",  # DE
    "A1RKKUPIHCS9HS",  # NL
    "A1805IZSGTT6HS",  # IT
    "AMEN7PMS3EDWL",   # BE/SE
]

# 10 SKU a exclure du feed v5 (a traiter manuellement via Seller Central)
SKIP_SKUS = {
    "I0735", "I0723", "I0628", "I0205",  # conflit ASIN TEA
    "C0184", "I0669",                      # brand registry only
}

# Colonnes lies aux variations a NEUTRALISER (passe en SKU simples)
VARIATION_FIELDS_TO_CLEAR = {
    "parent_sku",
    "parent_child",
    "relationship_type",
    "variation_theme",
    "size_name",
}

# Defaults nouveaux champs obligatoires (the - pas de batterie, pas dangereux)
DEFAULTS = {
    "batteries_required": "No",
    "are_batteries_included": "No",
    "battery_weight": "",
    "battery_weight_unit_of_measure": "",
    "battery_cell_composition": "",
    "lithium_battery_weight": "",
    "lithium_battery_weight_unit_of_measure": "",
    "lithium_battery_energy_content": "",
    "lithium_battery_energy_content_unit_of_measure": "",
    "lithium_battery_packaging": "",
    "number_of_lithium_ion_cells": "",
    "number_of_lithium_metal_cells": "",
    "hazmat_united_nations_regulatory_id": "",
    "contains_liquid_contents": "No",
    "is_heat_sensitive": "No",
    "contains_food_or_beverage": "Yes",
    "temperature_rating": "Ambiante : Température ambiante",
    "each_unit_count": "1",
    "item_volume": "",
    "item_volume_unit_of_measure": "",
    "safety_data_sheet_url": "",
    "fulfillment_availability#1.fulfillment_channel_code": "DEFAULT",
    "fulfillment_availability#1.is_inventory_available": "true",
    "fulfillment_availability#1.lead_time_to_ship_max_days": "3",
    "item_length_unit_of_measure": "centimeters",
    "item_width_unit_of_measure": "centimeters",
    "item_height_unit_of_measure": "centimeters",
}

SMART_QUOTE_MAP = str.maketrans({
    "‘": "'", "’": "'",
    "“": '"', "”": '"',
    "–": "-", "—": "-",
    "…": "...",
    " ": " ",
    " ": " ",
    " ": " ",
    "​": "",
    "­": "",
    "﻿": "",
})

def is_emoji_or_symbol(ch):
    cp = ord(ch)
    if 0x2300 <= cp <= 0x27BF: return True
    if 0x1F000 <= cp <= 0x1FFFF: return True
    if 0xFE00 <= cp <= 0xFE0F: return True
    if 0x200C <= cp <= 0x200F: return True
    return False

def clean_text(v):
    if v is None:
        return ""
    s = str(v)
    s = s.translate(SMART_QUOTE_MAP)
    s = "".join("" if is_emoji_or_symbol(c) else c for c in s)
    s = unicodedata.normalize("NFC", s)
    s = s.replace("\t", " ").replace("\r", " ").replace("\n", " ")
    while "  " in s:
        s = s.replace("  ", " ")
    return s.strip()

def to_cp1252_safe(s):
    try:
        s.encode("cp1252")
        return s
    except UnicodeEncodeError:
        out = []
        for ch in s:
            try:
                ch.encode("cp1252")
                out.append(ch)
            except UnicodeEncodeError:
                nf = unicodedata.normalize("NFKD", ch)
                ascii_part = nf.encode("ascii", "ignore").decode("ascii")
                out.append(ascii_part if ascii_part else "")
        return "".join(out)

def truncate_flavor(s, max_len=50):
    """Tronque flavor_name a 50 chars. Strategie :
    1. Si '<= 50, on garde.
    2. Si contient ' - ', prend la partie apres le dernier ' - ' (souvent le nom commercial).
    3. Sinon truncate brutal au dernier mot avant 50.
    """
    if not s or len(s) <= max_len:
        return s
    if " - " in s:
        tail = s.rsplit(" - ", 1)[1].strip()
        if 0 < len(tail) <= max_len:
            return tail
    # truncate au mot
    cut = s[:max_len]
    if " " in cut:
        cut = cut.rsplit(" ", 1)[0]
    return cut.strip()

# === Load NEW template ===
shutil.copy(NEW_TEMPLATE, OUT_XLSX)
wb_out = load_workbook(OUT_XLSX, keep_vba=False)
ws_mod = wb_out['Modèle']

new_headers = []
for c in next(ws_mod.iter_rows(min_row=3, max_row=3, values_only=True)):
    new_headers.append(c if c else "")
while new_headers and not new_headers[-1]:
    new_headers.pop()

print(f"New template : {len(new_headers)} colonnes")

hidx = {h: i + 1 for i, h in enumerate(new_headers)}

# === Load OLD data ===
wb_old = openpyxl.load_workbook(OLD, read_only=True, data_only=True)
ws_old = wb_old['Amazon FBA - FR']
old_rows = list(ws_old.iter_rows(values_only=True))
old_headers = [h or "" for h in old_rows[0]]
data_start = 2

old_idx = {h: i for i, h in enumerate(old_headers) if h}

RENAME = {
    "quantity": "fulfillment_availability#1.quantity",
}

PRICE_OLD_COL = "standard_price" if "standard_price" in old_idx else None

n_skus = 0
n_skipped_blocked = 0
skipped_empty = 0
sku_weight_fixed = []
sku_pkg_weight_fixed = []
sku_flavor_truncated = []
sku_was_variation = []

# Wipe old data rows in template (keep L1, L2, L3 only)
for r in range(4, ws_mod.max_row + 1):
    for c in range(1, len(new_headers) + 1):
        ws_mod.cell(row=r, column=c, value=None)

out_row = 4
for row_values in old_rows[data_start:]:
    if not row_values:
        skipped_empty += 1
        continue
    sku_v = row_values[old_idx.get("item_sku", 1)] if old_idx.get("item_sku") is not None else None
    if not sku_v:
        skipped_empty += 1
        continue
    sku = clean_text(sku_v)

    # Skip SKUs bloques (TEA conflit + brand registry)
    if sku in SKIP_SKUS:
        n_skipped_blocked += 1
        continue

    n_skus += 1

    out_cells = [""] * len(new_headers)
    for new_h, new_pos in hidx.items():
        if new_h in old_idx:
            v = row_values[old_idx[new_h]]
            out_cells[new_pos - 1] = clean_text(v)

    for old_h, new_h in RENAME.items():
        if old_h in old_idx and new_h in hidx:
            v = row_values[old_idx[old_h]]
            out_cells[hidx[new_h] - 1] = clean_text(v)

    # NEUTRALISER les champs de variation -> SKU simples
    was_var = False
    for vf in VARIATION_FIELDS_TO_CLEAR:
        if vf in hidx:
            cur = out_cells[hidx[vf] - 1]
            if cur:
                was_var = True
            out_cells[hidx[vf] - 1] = ""
    if was_var:
        sku_was_variation.append(sku)

    # Defaults nouveaux champs obligatoires
    for k, default_v in DEFAULTS.items():
        if k in hidx:
            cur = out_cells[hidx[k] - 1]
            if not cur:
                out_cells[hidx[k] - 1] = default_v

    # Prix TTC
    price_ht = None
    if PRICE_OLD_COL:
        raw_price = row_values[old_idx[PRICE_OLD_COL]]
        try:
            price_ht = float(raw_price)
        except (TypeError, ValueError):
            price_ht = None
    if price_ht is not None:
        price_ttc = round(price_ht * TVA_FR_THE, 2)
        if "list_price_with_tax" in hidx:
            out_cells[hidx["list_price_with_tax"] - 1] = f"{price_ttc:.2f}"
        for mp in MARKETPLACES:
            k = f"purchasable_offer[marketplace_id={mp}]#1.our_price#1.schedule#1.value_with_tax"
            if k in hidx:
                out_cells[hidx[k] - 1] = f"{price_ttc:.2f}"

    # item_weight 2 decimales
    if "item_weight" in hidx:
        v = out_cells[hidx["item_weight"] - 1]
        if v:
            try:
                f = float(v)
                v2 = f"{round(f, 2)}"
                out_cells[hidx["item_weight"] - 1] = v2
                if abs(f - round(f, 2)) > 1e-9:
                    sku_weight_fixed.append((sku, v, v2))
            except ValueError:
                pass

    # package_weight 2 decimales
    if "package_weight" in hidx:
        v = out_cells[hidx["package_weight"] - 1]
        if v:
            try:
                f = float(v)
                v2 = f"{round(f, 2)}"
                out_cells[hidx["package_weight"] - 1] = v2
                if abs(f - round(f, 2)) > 1e-9:
                    sku_pkg_weight_fixed.append((sku, v, v2))
            except ValueError:
                pass

    # flavor_name <= 50 chars
    if "flavor_name" in hidx:
        v = out_cells[hidx["flavor_name"] - 1]
        if v and len(v) > 50:
            v2 = truncate_flavor(v, 50)
            out_cells[hidx["flavor_name"] - 1] = v2
            sku_flavor_truncated.append((sku, v, v2))

    # feed_product_type = grocery
    if "feed_product_type" in hidx:
        out_cells[hidx["feed_product_type"] - 1] = "grocery"

    # Write
    for col_i, val in enumerate(out_cells, 1):
        ws_mod.cell(row=out_row, column=col_i, value=val)
    out_row += 1

print(f"\nSKU exportes : {n_skus}")
print(f"SKU skippes (bloques) : {n_skipped_blocked} ({sorted(SKIP_SKUS)})")
print(f"Lignes vides ignorees : {skipped_empty}")
print(f"SKU passes de variation a single : {len(sku_was_variation)}")
print(f"SKU item_weight arrondi : {len(sku_weight_fixed)}")
print(f"SKU package_weight arrondi : {len(sku_pkg_weight_fixed)}")
print(f"SKU flavor_name tronque : {len(sku_flavor_truncated)}")
for s, old, new in sku_flavor_truncated[:8]:
    print(f"  {s}: {old!r} -> {new!r}")

wb_out.save(OUT_XLSX)
print(f"\nXLSX sauvegarde : {OUT_XLSX}")

# === Export TXT CP1252 ===
wb2 = load_workbook(OUT_XLSX, data_only=True)
ws2 = wb2['Modèle']

n_lines = 0
n_replaced = 0
with open(OUT_TXT, "w", encoding="cp1252", errors="replace", newline="") as f:
    for row in ws2.iter_rows(values_only=True):
        cells = []
        for v in row:
            if v is None:
                cells.append("")
            else:
                if isinstance(v, (int, float)):
                    s = str(v)
                else:
                    s = clean_text(v)
                s_safe = to_cp1252_safe(s)
                if s_safe != s:
                    n_replaced += 1
                s_safe = s_safe.replace("\t", " ").replace("\r", " ").replace("\n", " ")
                cells.append(s_safe)
        while cells and cells[-1] == "":
            cells.pop()
        f.write("\t".join(cells))
        f.write("\r\n")
        n_lines += 1

print(f"TXT sauvegarde : {OUT_TXT}")
print(f"Lignes ecrites : {n_lines}")
print(f"Caracteres non Latin-1 remplaces : {n_replaced}")

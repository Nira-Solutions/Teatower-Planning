"""
Forecast échéancier Teatower — agent Compta.

Sources :
- Google Sheet Nicolas (onglets Échéancier + Budget Q1/Q2/Q3/Q4 2026)
- Odoo : account.move in_invoice non payées + historique fournisseurs récurrents

Sortie :
- Onglet "Forecast Odoo" dans le même Google Sheet :
  * Vue hebdo (WEEK XX) avec OUT prévus venant d'Odoo + Budget Qx prepaid
  * Comparaison avec l'Échéancier manuel de Nicolas → écart semaine par semaine
  * Détail ligne-à-ligne en bas (date, bénéficiaire, montant, source, référence)
"""

import argparse
import sys
import urllib.request
import xmlrpc.client
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

try:
    import gspread
    from google.oauth2.service_account import Credentials
    _HAS_GSPREAD = True
except ImportError:
    _HAS_GSPREAD = False

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─── Config ───────────────────────────────────────────────────────────────

SA_KEY = Path(__file__).parent / "google_service_account.json"
SHEET_ID = "1tpQQ5vTr5ekQesJKmkJi86cq9dG7sIFZ"
SHEET_XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
LOCAL_SNAPSHOT = Path(__file__).parent / "paiements_nicolas.xlsx"
LOCAL_FORECAST = Path(__file__).parent / "forecast_odoo.xlsx"
FORECAST_TAB = "Forecast Odoo"
ECHEANCIER_TAB = "Échéancier"
BUDGET_TABS = ["Budget Q1 2026", "Budget Q2 2026", "Budget Q3 2026", "Budget Q4 2026"]

ODOO_URL = "https://tea-tree.odoo.com"
ODOO_DB = "tsc-be-tea-tree-main-18515272"
ODOO_USER = "nicolas.raes@teatower.com"
ODOO_PWD = "Teatower123"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# ─── Google Sheets ────────────────────────────────────────────────────────

def gs_client():
    if not SA_KEY.exists():
        sys.exit(f"❌ Clé service account introuvable : {SA_KEY}\n"
                 f"   Voir compta/SETUP_GOOGLE_SHEETS.md")
    creds = Credentials.from_service_account_file(str(SA_KEY), scopes=SCOPES)
    return gspread.authorize(creds)


def test_access():
    gc = gs_client()
    sh = gc.open_by_key(SHEET_ID)
    tabs = [ws.title for ws in sh.worksheets()]
    print(f"✓ Accès OK — {len(tabs)} onglets : {tabs}")
    return tabs


def read_echeancier_weeks(sh):
    """Extrait la structure semaine de l'onglet Échéancier : WEEK XX → total OUT manuel.
    Permet de comparer avec Odoo."""
    try:
        ws = sh.worksheet(ECHEANCIER_TAB)
    except gspread.WorksheetNotFound:
        return {}
    rows = ws.get_all_values()
    weeks = defaultdict(lambda: {"out": 0.0, "in": 0.0, "items": []})
    # Col A = libellé semaine, col G (~index 6) = OUT, col L (~index 11) = IN
    for r in rows:
        if not r or len(r) < 12:
            continue
        week_label = (r[0] or "").strip()
        if not week_label.upper().startswith("WEEK"):
            continue
        try:
            out = float((r[6] or "0").replace(",", ".").replace(" ", "").replace("€", ""))
        except ValueError:
            out = 0.0
        try:
            inv = float((r[11] or "0").replace(",", ".").replace(" ", "").replace("€", ""))
        except ValueError:
            inv = 0.0
        label = r[1] if len(r) > 1 else ""
        weeks[week_label]["out"] += out
        weeks[week_label]["in"] += inv
        if label and out:
            weeks[week_label]["items"].append((label, out))
    return dict(weeks)


def read_budget_prepaid(sh):
    """Extrait les lignes Budget Qx avec payment terms 'prepaid' — à anticiper."""
    prepaid = []
    for tab in BUDGET_TABS:
        try:
            ws = sh.worksheet(tab)
        except gspread.WorksheetNotFound:
            continue
        for row in ws.get_all_records():
            terms = str(row.get("Payement terms", "") or row.get("Payment terms", "")).lower()
            cost = row.get("Coût") or row.get("Cost") or 0
            try:
                cost = float(str(cost).replace(",", ".").replace(" ", ""))
            except (ValueError, TypeError):
                cost = 0.0
            if cost and "prepaid" in terms:
                prepaid.append({
                    "quarter": tab,
                    "type": row.get("Type", ""),
                    "fournisseur": row.get("Fournisseur", "") or row.get("détail", "") or row.get("detail", ""),
                    "montant": cost,
                    "detail": row.get("détail", "") or row.get("detail", ""),
                })
    return prepaid

# ─── Odoo ─────────────────────────────────────────────────────────────────

def odoo_call():
    common = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/common")
    uid = common.authenticate(ODOO_DB, ODOO_USER, ODOO_PWD, {})
    m = xmlrpc.client.ServerProxy(f"{ODOO_URL}/xmlrpc/2/object")
    def call(model, method, args, kw=None):
        return m.execute_kw(ODOO_DB, uid, ODOO_PWD, model, method, args, kw or {})
    return call


def read_odoo_dues(call):
    moves = call("account.move", "search_read", [[
        ("move_type", "=", "in_invoice"),
        ("state", "=", "posted"),
        ("payment_state", "in", ["not_paid", "partial", "in_payment"]),
    ]], {"fields": [
        "name", "partner_id", "invoice_date", "invoice_date_due",
        "amount_total", "amount_residual", "payment_state", "ref",
    ], "order": "invoice_date_due asc"})
    return moves


def detect_recurrents_odoo(call, months_back=6):
    since = (datetime.today() - timedelta(days=30 * months_back)).strftime("%Y-%m-%d")
    moves = call("account.move", "search_read", [[
        ("move_type", "=", "in_invoice"),
        ("state", "=", "posted"),
        ("invoice_date", ">=", since),
    ]], {"fields": ["partner_id", "invoice_date", "amount_total"]})

    by_partner = defaultdict(list)
    for m in moves:
        by_partner[m["partner_id"][0]].append(m)

    # Factures déjà ouvertes : ne pas doubler le récurrent si déjà présent dans dues
    recurrents = []
    for pid, ms in by_partner.items():
        if len(ms) < 3:
            continue
        dates = sorted(datetime.strptime(x["invoice_date"], "%Y-%m-%d") for x in ms)
        gaps = [(dates[i+1] - dates[i]).days for i in range(len(dates)-1)]
        avg_gap = sum(gaps) / len(gaps)
        if 25 <= avg_gap <= 35:
            freq, period_days = "mensuel", 30
        elif 80 <= avg_gap <= 100:
            freq, period_days = "trimestriel", 90
        else:
            continue
        avg_amount = sum(x["amount_total"] for x in ms) / len(ms)
        next_date = dates[-1] + timedelta(days=round(avg_gap))
        # Ne projeter que 1 ou 2 prochaines échéances (3 mois max)
        projections = []
        d = next_date
        for _ in range(3):
            if (d - datetime.today()).days > 90:
                break
            projections.append(d)
            d = d + timedelta(days=period_days)
        for d in projections:
            recurrents.append({
                "partner_id": pid,
                "partner_name": ms[0]["partner_id"][1],
                "avg_amount": round(avg_amount, 2),
                "frequence": freq,
                "next_date": d.strftime("%Y-%m-%d"),
            })
    return recurrents

# ─── Forecast builder ─────────────────────────────────────────────────────

def _try_date(s):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except (ValueError, TypeError):
            pass
    return None


def week_label(dt):
    y, w, _ = dt.isocalendar()
    return f"WEEK {w}"


def build_rows(odoo_dues, odoo_recs, budget_prepaid):
    rows = []
    for m in odoo_dues:
        d = _try_date(m.get("invoice_date_due") or m.get("invoice_date"))
        if not d:
            continue
        rows.append({
            "date": d,
            "week": week_label(d),
            "beneficiaire": m["partner_id"][1] if m.get("partner_id") else "?",
            "montant": m.get("amount_residual", m.get("amount_total", 0)),
            "source": "Odoo — facture posted",
            "reference": m.get("name") or m.get("ref") or "",
            "statut": m.get("payment_state", ""),
        })
    for r in odoo_recs:
        d = _try_date(r["next_date"])
        rows.append({
            "date": d,
            "week": week_label(d),
            "beneficiaire": r["partner_name"],
            "montant": r["avg_amount"],
            "source": f"Odoo — récurrent {r['frequence']} projeté",
            "reference": "",
            "statut": "projeté",
        })
    # Budget prepaid : sans date précise, on les met tous au début du trimestre
    quarter_starts = {
        "Budget Q1 2026": datetime(2026, 1, 1),
        "Budget Q2 2026": datetime(2026, 4, 1),
        "Budget Q3 2026": datetime(2026, 7, 1),
        "Budget Q4 2026": datetime(2026, 10, 1),
    }
    for b in budget_prepaid:
        d = quarter_starts.get(b["quarter"])
        if not d or d < datetime.today() - timedelta(days=15):
            continue  # passé
        rows.append({
            "date": d,
            "week": week_label(d),
            "beneficiaire": b["fournisseur"] or b["detail"] or b["type"],
            "montant": b["montant"],
            "source": f"Budget {b['quarter'].replace('Budget ', '')} prepaid",
            "reference": b["detail"],
            "statut": "planifié",
        })
    rows.sort(key=lambda x: x["date"])
    return rows


def write_forecast(gc, rows, echeancier_weeks):
    sh = gc.open_by_key(SHEET_ID)
    try:
        ws = sh.worksheet(FORECAST_TAB)
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=FORECAST_TAB, rows=max(len(rows)+60, 200), cols=10)

    now = datetime.today()

    # Agrégats par semaine
    by_week = defaultdict(lambda: {"total": 0.0, "items": []})
    for r in rows:
        by_week[r["week"]]["total"] += r["montant"]
        by_week[r["week"]]["items"].append(r)

    # Tri semaines par première date
    week_order = sorted(by_week.keys(),
                        key=lambda w: min(r["date"] for r in by_week[w]["items"]))

    # Totaux rapides
    j7 = sum(r["montant"] for r in rows if r["date"] <= now + timedelta(days=7))
    j30 = sum(r["montant"] for r in rows if r["date"] <= now + timedelta(days=30))
    total = sum(r["montant"] for r in rows)

    block = []
    block.append([f"Forecast Odoo — généré {now.strftime('%Y-%m-%d %H:%M')} ({len(rows)} lignes)"])
    block.append([f"À payer J+7 : {j7:,.2f} EUR   |   J+30 : {j30:,.2f} EUR   |   Total : {total:,.2f} EUR"])
    block.append([])

    # Vue hebdo + comparaison Échéancier manuel
    block.append(["VUE HEBDOMADAIRE (comparaison avec Échéancier manuel)"])
    block.append(["Semaine", "OUT Odoo prévu", "OUT Échéancier manuel", "Écart (Odoo - Manuel)", "Nb lignes Odoo"])
    for w in week_order:
        out_odoo = by_week[w]["total"]
        manual = echeancier_weeks.get(w, {}).get("out", 0.0)
        # l'échéancier manuel stocke les OUT en négatif
        manual_abs = abs(manual)
        ecart = out_odoo - manual_abs
        block.append([w,
                      f"{out_odoo:,.2f}",
                      f"{manual_abs:,.2f}" if manual_abs else "—",
                      f"{ecart:+,.2f}" if manual_abs else "n/a",
                      len(by_week[w]["items"])])
    block.append([])

    # Détail ligne à ligne
    block.append(["DÉTAIL LIGNE À LIGNE"])
    block.append(["Date échéance", "Semaine", "Bénéficiaire", "Montant", "Source", "Référence", "Statut"])
    for r in rows:
        block.append([
            r["date"].strftime("%Y-%m-%d"),
            r["week"],
            r["beneficiaire"],
            r["montant"],
            r["source"],
            r["reference"],
            r["statut"],
        ])

    ws.update("A1", block, value_input_option="USER_ENTERED")

    # Mise en forme minimale : bold headers, colonne montant format EUR
    ws.format("A1", {"textFormat": {"bold": True, "fontSize": 12}})
    ws.format("A4", {"textFormat": {"bold": True}, "backgroundColor": {"red": 0.85, "green": 0.92, "blue": 0.98}})
    title_rows = [i+1 for i, r in enumerate(block) if r and isinstance(r[0], str) and r[0].startswith(("VUE", "DÉTAIL"))]
    for tr in title_rows:
        ws.format(f"A{tr}", {"textFormat": {"bold": True}, "backgroundColor": {"red": 0.85, "green": 0.92, "blue": 0.98}})

    print(f"✓ '{FORECAST_TAB}' : {len(rows)} lignes · J+7 {j7:,.2f} · J+30 {j30:,.2f} · Total {total:,.2f}")

# ─── Main ─────────────────────────────────────────────────────────────────

def print_dry_run(rows, weeks):
    now = datetime.today()
    overdue = sum(r["montant"] for r in rows if r["date"] < now)
    n_overdue = sum(1 for r in rows if r["date"] < now)
    j7 = sum(r["montant"] for r in rows if now <= r["date"] <= now + timedelta(days=7))
    j30 = sum(r["montant"] for r in rows if now <= r["date"] <= now + timedelta(days=30))
    total = sum(r["montant"] for r in rows)

    print()
    print("=" * 80)
    print(f"DRY-RUN — aucune écriture")
    print("=" * 80)
    print(f"⚠ En retard (date échéance passée) : {overdue:,.2f} EUR sur {n_overdue} factures")
    print(f"À payer — J+7 : {j7:,.2f} EUR  |  J+30 : {j30:,.2f} EUR")
    print(f"Total forecast : {total:,.2f} EUR ({len(rows)} lignes)")
    print()

    # Vue hebdo
    by_week = defaultdict(lambda: {"total": 0.0, "n": 0})
    for r in rows:
        by_week[r["week"]]["total"] += r["montant"]
        by_week[r["week"]]["n"] += 1
    week_order = sorted(by_week.keys(),
                        key=lambda w: min(r["date"] for r in rows if r["week"] == w))

    print(f"{'Semaine':<10} {'Odoo prévu':>14} {'Manuel':>14} {'Écart':>14} {'Lignes':>8}")
    print("-" * 64)
    for w in week_order[:20]:
        out_odoo = by_week[w]["total"]
        manual = abs(weeks.get(w, {}).get("out", 0.0))
        ecart = out_odoo - manual
        manual_str = f"{manual:,.2f}" if manual else "—"
        ecart_str = f"{ecart:+,.2f}" if manual else "n/a"
        print(f"{w:<10} {out_odoo:>14,.2f} {manual_str:>14} {ecart_str:>14} {by_week[w]['n']:>8}")

    print()
    print(f"Détail — 15 premières échéances :")
    print("-" * 80)
    for r in rows[:15]:
        print(f"  {r['date'].strftime('%Y-%m-%d')} | {r['beneficiaire'][:30]:<30} | {r['montant']:>10,.2f} | {r['source']}")
    if len(rows) > 15:
        print(f"  … +{len(rows)-15} autres lignes")
    print()


# ─── Mode local (sans Google Cloud) ──────────────────────────────────────

def download_snapshot():
    print(f"→ Téléchargement du Sheet via URL publique…")
    urllib.request.urlretrieve(SHEET_XLSX_URL, LOCAL_SNAPSHOT)
    print(f"  Sauvegardé : {LOCAL_SNAPSHOT}")


def read_echeancier_weeks_local(wb):
    # Fuzzy match sur le nom (accents peuvent différer selon l'encodage)
    tab = None
    for name in wb.sheetnames:
        if "ch" in name.lower() and "ancier" in name.lower():
            tab = name
            break
    if not tab:
        return {}
    ws = wb[tab]
    weeks = defaultdict(lambda: {"out": 0.0, "in": 0.0, "items": []})
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 12:
            continue
        week_label = str(row[0] or "").strip()
        if not week_label.upper().startswith("WEEK"):
            continue
        try:
            out = float(str(row[6] or 0).replace(",", ".").replace(" ", ""))
        except (ValueError, TypeError):
            out = 0.0
        try:
            inv = float(str(row[11] or 0).replace(",", ".").replace(" ", ""))
        except (ValueError, TypeError):
            inv = 0.0
        weeks[week_label]["out"] += out
        weeks[week_label]["in"] += inv
    return dict(weeks)


def read_budget_prepaid_local(wb):
    prepaid = []
    for tab in BUDGET_TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h or "").strip() for h in rows[0]]
        idx = {h.lower(): i for i, h in enumerate(headers)}
        for row in rows[1:]:
            if not row or not any(row):
                continue
            def g(key):
                for k in idx:
                    if key in k:
                        return row[idx[k]]
                return None
            terms = str(g("payement") or g("payment") or "").lower()
            cost = g("coût") or g("cout") or g("cost") or 0
            try:
                cost = float(str(cost).replace(",", ".").replace(" ", ""))
            except (ValueError, TypeError):
                cost = 0.0
            if cost and "prepaid" in terms:
                prepaid.append({
                    "quarter": tab,
                    "type": g("type") or "",
                    "fournisseur": g("fournisseur") or g("détail") or g("detail") or "",
                    "montant": cost,
                    "detail": g("détail") or g("detail") or "",
                })
    return prepaid


def write_local_xlsx(rows, weeks):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = FORECAST_TAB

    now = datetime.today()
    j7 = sum(r["montant"] for r in rows if r["date"] <= now + timedelta(days=7))
    j30 = sum(r["montant"] for r in rows if r["date"] <= now + timedelta(days=30))
    total = sum(r["montant"] for r in rows)

    bold = Font(bold=True)
    blue = PatternFill("solid", fgColor="D9E7F5")
    title = Font(bold=True, size=13)

    ws["A1"] = f"Forecast Odoo — généré {now.strftime('%Y-%m-%d %H:%M')} ({len(rows)} lignes)"
    ws["A1"].font = title
    ws["A2"] = f"À payer J+7 : {j7:,.2f} EUR   |   J+30 : {j30:,.2f} EUR   |   Total : {total:,.2f} EUR"
    ws["A2"].font = bold

    # Vue hebdo
    r = 4
    ws.cell(r, 1, "VUE HEBDOMADAIRE (comparaison avec Échéancier manuel)").font = bold
    ws.cell(r, 1).fill = blue
    r += 1
    headers = ["Semaine", "OUT Odoo prévu", "OUT Échéancier manuel", "Écart", "Nb lignes Odoo"]
    for i, h in enumerate(headers):
        c = ws.cell(r, i+1, h); c.font = bold; c.fill = blue
    r += 1

    by_week = defaultdict(lambda: {"total": 0.0, "items": []})
    for row in rows:
        by_week[row["week"]]["total"] += row["montant"]
        by_week[row["week"]]["items"].append(row)
    week_order = sorted(by_week.keys(),
                        key=lambda w: min(row["date"] for row in by_week[w]["items"]))

    for w in week_order:
        out_odoo = by_week[w]["total"]
        manual = abs(weeks.get(w, {}).get("out", 0.0))
        ecart = out_odoo - manual if manual else None
        ws.cell(r, 1, w)
        ws.cell(r, 2, round(out_odoo, 2))
        ws.cell(r, 3, round(manual, 2) if manual else "—")
        ws.cell(r, 4, round(ecart, 2) if ecart is not None else "n/a")
        ws.cell(r, 5, len(by_week[w]["items"]))
        r += 1

    r += 2
    ws.cell(r, 1, "DÉTAIL LIGNE À LIGNE").font = bold
    ws.cell(r, 1).fill = blue
    r += 1
    detail_headers = ["Date échéance", "Semaine", "Bénéficiaire", "Montant", "Source", "Référence", "Statut"]
    for i, h in enumerate(detail_headers):
        c = ws.cell(r, i+1, h); c.font = bold; c.fill = blue
    r += 1
    for row in rows:
        ws.cell(r, 1, row["date"].strftime("%Y-%m-%d"))
        ws.cell(r, 2, row["week"])
        ws.cell(r, 3, row["beneficiaire"])
        ws.cell(r, 4, round(row["montant"], 2))
        ws.cell(r, 5, row["source"])
        ws.cell(r, 6, row["reference"])
        ws.cell(r, 7, row["statut"])
        r += 1

    # Auto-size cols
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value), default=12)
        ws.column_dimensions[col[0].column_letter].width = min(length + 2, 45)

    wb.save(LOCAL_FORECAST)
    print(f"✓ Fichier généré : {LOCAL_FORECAST}")
    print(f"  → Ouvre-le, copie le contenu, colle dans un nouvel onglet 'Forecast Odoo' de ton Google Sheet")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--test", action="store_true", help="Teste l'accès Google Sheets uniquement")
    p.add_argument("--dry-run", action="store_true", help="Affiche le résultat sans écrire")
    p.add_argument("--local", action="store_true",
                   help="Pas besoin de service account. Télécharge le Sheet public, écrit un .xlsx local à coller à la main.")
    args = p.parse_args()

    if args.local:
        download_snapshot()
        wb = openpyxl.load_workbook(LOCAL_SNAPSHOT, data_only=True)
        weeks = read_echeancier_weeks_local(wb)
        budget = read_budget_prepaid_local(wb)
        print(f"  Échéancier : {len(weeks)} semaines, Budget prepaid : {len(budget)} lignes")

        print("→ Lecture Odoo…")
        call = odoo_call()
        dues = read_odoo_dues(call)
        recs = detect_recurrents_odoo(call)
        print(f"  {len(dues)} factures ouvertes, {len(recs)} récurrents projetés")

        rows = build_rows(dues, recs, budget)

        if args.dry_run:
            print_dry_run(rows, weeks)
            return

        write_local_xlsx(rows, weeks)
        return

    if not _HAS_GSPREAD:
        sys.exit("❌ gspread non installé et pas de mode --local demandé.\n"
                 "   Soit : pip install gspread google-auth\n"
                 "   Soit : relance avec --local (sans Google Cloud)")

    if args.test:
        test_access()
        return

    print("→ Lecture Google Sheet…")
    gc = gs_client()
    sh = gc.open_by_key(SHEET_ID)
    weeks = read_echeancier_weeks(sh)
    budget = read_budget_prepaid(sh)
    print(f"  Échéancier : {len(weeks)} semaines trackées")
    print(f"  Budget prepaid : {len(budget)} lignes")

    print("→ Lecture Odoo…")
    call = odoo_call()
    dues = read_odoo_dues(call)
    recs = detect_recurrents_odoo(call)
    print(f"  {len(dues)} factures fournisseurs ouvertes, {len(recs)} récurrents projetés")

    print("→ Génération…")
    rows = build_rows(dues, recs, budget)

    if args.dry_run:
        print_dry_run(rows, weeks)
        return

    print("→ Écriture…")
    write_forecast(gc, rows, weeks)


if __name__ == "__main__":
    main()

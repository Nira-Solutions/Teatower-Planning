#!/usr/bin/env python3
"""Génère Fiches_produits_HC250_2026-06-19.xlsx — lecture seule Odoo + Shopify."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Fiches HC250"

# ─── STYLES ───────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F3864")
HDR_FONT  = Font(color="FFFFFF", bold=True, size=11)
CAT_FILL  = PatternFill("solid", fgColor="BDD7EE")
CAT_FONT  = Font(bold=True, size=11)
PROD_FILL = PatternFill("solid", fgColor="E2EFDA")
HC_FILL   = PatternFill("solid", fgColor="FCE4D6")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
LABEL_FONT = Font(bold=True, size=10)
VAL_FONT   = Font(size=10)
WRAP   = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
thin   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(cell, fill=None, font=None, alignment=None):
    if fill:      cell.fill      = fill
    if font:      cell.font      = font
    if alignment: cell.alignment = alignment
    cell.border = BORDER

# ─── DONNÉES ──────────────────────────────────────────────────────────────────
products = [
    {
        "hc_code":        "HC250735",
        "hc_name":        "Peche de vigne_ BE-BIO-01 25 enveloppes",
        "hc_template_id": 4616,
        "hc_barcode":     "5413393814737",
        "hc_list_price":  "10,00",
        "hc_std_price":   "3,56",
        "hc_actif":       "Oui",
        "hc_categ":       "Horeca",
        "hc_etat":        "Fiche creee, prix/barcode OK — AUCUNE description/ingredients renseignes",
        "v0_code":        "V0735",
        "v0_name":        "Peche de vigne - BIO (100 g vrac)",
        "v0_template_id": 5117,
        "v0_barcode":     "5413393800525",
        "v0_list_price":  "9,43",
        "v0_std_price":   "2,28",
        "v0_categ":       "Rooibos",
        "famille":        "Rooibos",
        "type_the":       "Rooibos",
        "is_bio":         "Oui - BE-BIO-01",
        "origine":        "Afrique du Sud (Rooibos AOP)",
        "poids_v0":       "100 g vrac",
        "poids_hc":       "25 enveloppes (approx. 62,5 g a 2,5 g/env.)",
        "theine":         "Sans theine",
        "ingredients":    (
            "Rooibos vert* (appellation d'origine protegee), morceaux de pomme*, "
            "zestes de citron*, arome naturel de peche, soucis*, poivre rose*, "
            "arome naturel de vanille. "
            "*ingredient d'origine biologique"
        ),
        "allergenes":     "Aucun allergene majeur declare (verifier arome naturel avec fournisseur)",
        "temperature":    "90 degres C",
        "infusion":       "5 min",
        "dosage":         "Non precise sur fiche (standard : ~2,5 g pour 25 cl)",
        "moment":         "Au choix - toute la journee",
        "desc_courte":    (
            "L'infusion de rooibos Peche de vigne - BIO met en avant la douceur de la peche, "
            "soutenue par des notes legerement citronnees et epicees. "
            "Naturellement sans theine, ideale a tout moment de la journee."
        ),
        "desc_longue":    (
            "Rooibos vert AOP d'Afrique du Sud, certifie agriculture biologique BE-BIO-01. "
            "Notes dominantes : peche douce, legerement acidulee citron, epices douces "
            "(poivre rose, vanille). Sans theine - convient a toute la famille, "
            "du matin jusqu'au soir."
        ),
        "source_desc":    "Shopify I0735 (tmpl Odoo 4689)",
        "points_attention": (
            "- Rooibos = pas de theine, pas de cafeine -> communication sante sans restriction\n"
            "- Verifier allergenes arome naturel sur fiche fournisseur\n"
            "- standard_price HC = 3,56 EUR (coherent format Horeca 25 env.)"
        ),
    },
    {
        "hc_code":        "HC250626",
        "hc_name":        "Sencha_BE-BIO-01 - 25 enveloppes",
        "hc_template_id": 4608,
        "hc_barcode":     "5413393814546",
        "hc_list_price":  "10,00",
        "hc_std_price":   "3,56",
        "hc_actif":       "Oui",
        "hc_categ":       "Horeca",
        "hc_etat":        "Fiche creee, prix/barcode OK — AUCUNE description/ingredients renseignes",
        "v0_code":        "V0626",
        "v0_name":        "Sencha - BIO (100 g vrac)",
        "v0_template_id": 5077,
        "v0_barcode":     "5413393802161",
        "v0_list_price":  "9,43",
        "v0_std_price":   "1,55",
        "v0_categ":       "The Vert",
        "famille":        "The Vert",
        "type_the":       "The vert - Sencha",
        "is_bio":         "Oui - BE-BIO-01",
        "origine":        "Asie (Japon / Chine selon approvisionnement)",
        "poids_v0":       "100 g vrac",
        "poids_hc":       "25 enveloppes (approx. 37,5 g a 1,5 g/env.)",
        "theine":         "Avec theine",
        "ingredients":    "Ingredient : the vert*. *Produit issu de l'agriculture biologique.",
        "allergenes":     "Aucun allergene majeur declare",
        "temperature":    "70 degres C",
        "infusion":       "1 a 3 min",
        "dosage":         "Non precise (standard : ~2 g pour 20 cl)",
        "moment":         "Journee",
        "desc_courte":    (
            "Le Sencha - BIO est un the vert biologique emblematique, apprecie pour son profil "
            "vegetal et doux. Issu de l'agriculture biologique, il incarne la purete et "
            "l'authenticite des grands thes verts traditionnels."
        ),
        "desc_longue":    (
            "En infusion, ce sencha devoile une tasse fraiche et equilibree, aux notes vegetales "
            "delicates et a la douceur naturelle. Naturellement avec theine, il est ideal pour "
            "accompagner la journee. Certifie agriculture biologique BE-BIO-01."
        ),
        "source_desc":    "Shopify I0626 (tmpl Odoo 4657)",
        "points_attention": (
            "- The vert -> temperature basse 70 degres C OBLIGATOIRE (amertume si depassee)\n"
            "- standard_price HC = 3,56 EUR vs V0 = 1,55 EUR (format Horeca justifie l'ecart)\n"
            "- Mentions BIO a reporter integralement sur etiquette Horeca (INCO UE)"
        ),
    },
    {
        "hc_code":        "HC250121",
        "hc_name":        "Lady Dodo 25 enveloppes",
        "hc_template_id": 4599,
        "hc_barcode":     "5413393814638",
        "hc_list_price":  "10,00",
        "hc_std_price":   "3,56",
        "hc_actif":       "Oui",
        "hc_categ":       "Horeca",
        "hc_etat":        "Fiche creee, prix/barcode OK — AUCUNE description/ingredients renseignes",
        "v0_code":        "V0121",
        "v0_name":        "Lady Dodo (80 g vrac)",
        "v0_template_id": 5036,
        "v0_barcode":     "5413393801188",
        "v0_list_price":  "9,43",
        "v0_std_price":   "2,11",
        "v0_categ":       "Plantes",
        "famille":        "Infusion de plantes",
        "type_the":       "Infusion (sans theine)",
        "is_bio":         "Non",
        "origine":        "Melange Europe / zones temperees",
        "poids_v0":       "80 g vrac",
        "poids_hc":       "25 enveloppes (approx. 75 g a 3 g/env.)",
        "theine":         "Sans theine (0%)",
        "ingredients":    (
            "Graines de fenouil (20,5%), zestes d'orange, verveine, "
            "feuilles de mures douces, camomille (10,5%), arome naturel."
        ),
        "allergenes":     "Aucun allergene majeur declare — verifier arome naturel",
        "temperature":    "95-100 degres C (eau bien chaude pour liberer les huiles essentielles du fenouil)",
        "infusion":       "7 a 10 min",
        "dosage":         "1 cuillere a cafe bombee (3 g) pour 25 cl",
        "moment":         "30-45 min avant le coucher",
        "desc_courte":    (
            "Lady Dodo est l'infusion du soir par excellence. Fenouil, camomille, verveine et "
            "zestes d'orange forment un rituel doux et apaisant. "
            "Sans theine. La plus re-commandee de Teatower (note 4,64/5 - 72 avis)."
        ),
        "desc_longue":    (
            "En tasse, robe doree aux reflets ambres. Le fenouil apporte ses notes anisees et "
            "chaleureuses, la camomille distille sa douceur florale. Les zestes d'orange petillent "
            "discretement, la verveine apaise, et les feuilles de mures douces enveloppent "
            "l'ensemble. Chaque gorgee est une invitation au lacher-prise."
        ),
        "source_desc":    "Shopify I0121 (tmpl Odoo 4628)",
        "points_attention": (
            "- INFUSION (pas the) -> sans theine -> communication Horeca : tisane du soir / sommeil\n"
            "- Dosage 3 g/enveloppe : verifier grammage enveloppe HC (standard = 1,5-2 g/env.)\n"
            "- Fenouil 20,5% = premier ingredient -> mention obligatoire INCO etiquette\n"
            "- is_bio = Non sur fiche Odoo -> ne PAS ajouter mention BIO"
        ),
    },
    {
        "hc_code":        "HC250666",
        "hc_name":        "English Breakfast 25 enveloppes",
        "hc_template_id": 4612,
        "hc_barcode":     "5413393003018",
        "hc_list_price":  "10,00",
        "hc_std_price":   "0,00  *** ALERTE : prix achat manquant ***",
        "hc_actif":       "Oui",
        "hc_categ":       "Horeca",
        "hc_etat":        "ALERTE : standard_price = 0,00 EUR — prix achat manquant a corriger !",
        "v0_code":        "V0666",
        "v0_name":        "English Breakfast  (100 g vrac)",
        "v0_template_id": 5092,
        "v0_barcode":     "5413393802536",
        "v0_list_price":  "9,43",
        "v0_std_price":   "1,55",
        "v0_categ":       "The Noir",
        "famille":        "The Noir",
        "type_the":       "The noir - melange Assam / Ceylan / Inde",
        "is_bio":         "Non",
        "origine":        "Inde (Assam / Darjeeling selon blend)",
        "poids_v0":       "100 g vrac",
        "poids_hc":       "25 enveloppes (approx. 50 g a 2 g/env.)",
        "theine":         "Fort en theine",
        "ingredients":    "Ingredient : the noir d'Inde.",
        "allergenes":     "Aucun allergene majeur declare",
        "temperature":    "90 degres C",
        "infusion":       "3 a 4 min",
        "dosage":         "Non precise (standard : ~2 g pour 25 cl)",
        "moment":         "Matin - petit-dejeuner",
        "desc_courte":    (
            "L'English Breakfast est un the noir emblematique au caractere tannique et puissant. "
            "Intense et corse, il offre une infusion robuste et structuree, "
            "ideale pour accompagner le petit-dejeuner."
        ),
        "desc_longue":    (
            "Issu de the noir d'Inde, ce grand classique seduit par sa force aromatique et sa belle "
            "longueur en bouche. Naturellement fort en theine, il procure une energie franche et "
            "durable. Apprecie nature ou avec un nuage de lait selon les habitudes britanniques. "
            "Incontournable des cartes Horeca."
        ),
        "source_desc":    "Shopify I0666 (tmpl Odoo 4668)",
        "points_attention": (
            "- URGENT : standard_price = 0,00 EUR a corriger (les 3 autres HC sont a 3,56 EUR)\n"
            "- Barcode HC250666 (5413393003018) = format different des autres HC250 (5413393814xxx) — verifier\n"
            "- Origine 'Inde' a preciser sur etiquette INCO Horeca (Assam/Ceylan/blend)"
        ),
    },
]

SECTIONS = [
    ("IDENTIFICATION",              None,                None),
    ("Ref HC (nouveau code)",        "hc_code",           HC_FILL),
    ("Nom HC Odoo",                 "hc_name",           HC_FILL),
    ("Template ID Odoo HC",         "hc_template_id",    HC_FILL),
    ("Code-barres HC",              "hc_barcode",        HC_FILL),
    ("Prix vente HC HT (EUR)",      "hc_list_price",     HC_FILL),
    ("Prix achat HC (EUR)",         "hc_std_price",      HC_FILL),
    ("Actif Odoo",                  "hc_actif",          HC_FILL),
    ("Categorie Odoo HC",           "hc_categ",          HC_FILL),
    ("Etat fiche HC Odoo",          "hc_etat",           HC_FILL),
    ("",                            None,                None),
    ("REFERENCE SOURCE V0",         None,                None),
    ("Ref V0 (source ingrédients)", "v0_code",           PROD_FILL),
    ("Nom V0 Odoo",                 "v0_name",           PROD_FILL),
    ("Template ID Odoo V0",         "v0_template_id",    PROD_FILL),
    ("Code-barres V0",              "v0_barcode",        PROD_FILL),
    ("Prix vente V0 HT (EUR)",      "v0_list_price",     PROD_FILL),
    ("Prix achat V0 (EUR)",         "v0_std_price",      PROD_FILL),
    ("Categorie Odoo V0",           "v0_categ",          PROD_FILL),
    ("",                            None,                None),
    ("CARACTERISTIQUES PRODUIT",    None,                None),
    ("Famille / Gamme",             "famille",           None),
    ("Type de the / produit",       "type_the",          None),
    ("BIO / Certification",         "is_bio",            None),
    ("Origine",                     "origine",           None),
    ("Poids V0 (vrac)",             "poids_v0",          None),
    ("Format HC (Horeca)",          "poids_hc",          None),
    ("Theine",                      "theine",            None),
    ("",                            None,                None),
    ("COMPOSITION — INCO UE",       None,                None),
    ("Ingredients",                 "ingredients",       None),
    ("Allergenes",                  "allergenes",        None),
    ("",                            None,                None),
    ("CONSEILS DE PREPARATION",     None,                None),
    ("Temperature",                 "temperature",       None),
    ("Duree d infusion",            "infusion",          None),
    ("Dosage",                      "dosage",            None),
    ("Moment de consommation",      "moment",            None),
    ("",                            None,                None),
    ("CONTENUS MARKETING",          None,                None),
    ("Description courte (200 c.)", "desc_courte",       None),
    ("Description longue",          "desc_longue",       None),
    ("Source des descriptions",     "source_desc",       None),
    ("",                            None,                None),
    ("POINTS D ATTENTION",          None,                None),
    ("Points d attention",          "points_attention",  WARN_FILL),
]

HEADERS = ["Champ"] + [
    "{}\n<- {}".format(p["hc_code"], p["v0_code"]) for p in products
]

ws.append(HEADERS)
for col_idx, hdr in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.value = hdr
    style(cell, fill=HDR_FILL, font=HDR_FONT, alignment=CENTER)
ws.row_dimensions[1].height = 40

row = 2
for label, field, fill_override in SECTIONS:
    if field is None:
        ws.append([label] + [""] * len(products))
        cell = ws.cell(row=row, column=1)
        if label:
            cell.fill = CAT_FILL
            cell.font = CAT_FONT
            cell.alignment = CENTER
            ws.merge_cells(
                start_row=row, start_column=1,
                end_row=row, end_column=1 + len(products)
            )
            ws.row_dimensions[row].height = 22
        else:
            ws.row_dimensions[row].height = 6
        row += 1
        continue

    row_data = [label]
    for p in products:
        row_data.append(str(p.get(field, "")))
    ws.append(row_data)

    lc = ws.cell(row=row, column=1)
    lc.font = LABEL_FONT
    lc.alignment = WRAP
    lc.border = BORDER

    for col_idx in range(2, 2 + len(products)):
        vc = ws.cell(row=row, column=col_idx)
        vc.font = VAL_FONT
        vc.alignment = WRAP
        vc.border = BORDER
        if fill_override:
            vc.fill = fill_override

    max_lines = max(
        str(p.get(field, "")).count("\n") + 1 for p in products
    )
    ws.row_dimensions[row].height = max(18, min(120, max_lines * 16 + 6))
    row += 1

# ─── Onglet Résumé ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Resume")
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 55
ws2.column_dimensions["C"].width = 55

header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1F3864")

ws2.append(["Fichier", "Fiches produits HC250 — heritage V0", ""])
ws2.append(["Date generation", "2026-06-19", ""])
ws2.append(["Source", "Odoo tea-tree.odoo.com + Shopify descriptions I0xxx", ""])
ws2.append([])
ws2.append(["Correspondance", "Nom HC Odoo", "Statut fiche HC"])
for col in [1, 2, 3]:
    c = ws2.cell(row=5, column=col)
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.font = Font(bold=True, color="FFFFFF")
    c.alignment = CENTER

alt = [PatternFill("solid", fgColor="E2EFDA"), PatternFill("solid", fgColor="FCE4D6")]
for i, p in enumerate(products):
    ws2.append([
        "{} <- {}".format(p["hc_code"], p["v0_code"]),
        p["hc_name"],
        p["hc_etat"],
    ])
    for col in [1, 2, 3]:
        ws2.cell(row=6 + i, column=col).fill = alt[i % 2]

ws2.append([])
ws2.append(["POINTS D ATTENTION GLOBAUX"])
ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=11)

alertes = [
    ("URGENT", "HC250666 — standard_price = 0,00 EUR : a corriger immediatement (les 3 autres HC sont a 3,56 EUR)"),
    ("Info", "Aucune des 4 fiches HC n a de description / ingredients dans Odoo — ce fichier est la source pour remplissage"),
    ("Info", "Aucune des 4 fiches V0 n a de description dans Odoo non plus — source = Shopify I0xxx"),
    ("Verif", "HC250666 barcode 5413393003018 (format different des 3 autres HC250 5413393814xxx) — a contrôler"),
    ("Verif", "Lady Dodo (HC250121) — pas BIO, dosage 3 g/env. eleve vs standard 1,5-2 g -> confirmer avec fournisseur"),
]
for tag, msg in alertes:
    ws2.append([tag, msg])
    c_tag = ws2.cell(row=ws2.max_row, column=1)
    c_tag.font = Font(bold=True)
    if tag == "URGENT":
        c_tag.fill = PatternFill("solid", fgColor="FF0000")
        c_tag.font = Font(bold=True, color="FFFFFF")

ws2.row_dimensions[1].height = 22

# ─── Largeurs colonnes principale ─────────────────────────────────────────────
ws.column_dimensions["A"].width = 32
for col_idx in range(2, 2 + len(products)):
    ws.column_dimensions[get_column_letter(col_idx)].width = 48

ws.freeze_panes = "B2"

out = r"C:\Users\FlowUP\OneDrive\Teatower\Fiches_produits_HC250_2026-06-19.xlsx"
wb.save(out)
print("OK:", out)

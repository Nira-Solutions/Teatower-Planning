# -*- coding: utf-8 -*-
"""Génère les bons de commande GMS NOËL 2026 (INFUS + VRAC).

Base = « Tarif et Bon de Commande GMS - INFU/VRAC - MAJ Aout 2026.xlsx ».
Ajoute une section « Offre de Noël 2026 » ne contenant que C0195 Calendrier de l'avent.
Les autres coffrets Noël (C0196 Cachuète, C0197 Coffret découverte, C0198 boîte métal)
sont volontairement HORS périmètre GMS.

Prix : PVC 30,00 € TTC -> 28,30 € HTVA -> remise GMS 30 % -> 19,81 € net,
soit exactement le prix pratiqué sur le calendrier 2025 (150 pcs sur 157).
"""
import os
import sys
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

BASE = r"C:\Users\FlowUP\OneDrive\Teatower\Maj tarifaire Aout 2026"

# --- la ligne Noël, identique sur les deux bons ---
NOEL = {
    "A": "C0195",
    "B": "Calendrier de l'avent 2026",
    "C": "Coffret 24 infusettes",
    "D": "Coffret Noël",
    "E": "24 thés et infusions — « 24 moments magiques »",
    "H": 0.3,
    "I": 19.81,
    "J": 28.30,
    "K": 30.00,
    "L": 0.06,
    "O": "5413393004053",
}
TITRE_SECTION = "OFFRE DE NOËL 2026"

JOBS = [
    {
        "src": "Tarif et Bon de Commande GMS - INFU - MAJ Aout 2026.xlsx",
        "dst": "Tarif et Bon de Commande GMS NOEL 2026 - INFUS.xlsx",
        "sheet": "GMS NOEL 2026 INFUS",
        "titre": "TARIF & BON DE COMMANDE — GMS NOËL 2026 INFUSETTES",
        "last_data": 34,
    },
    {
        "src": "Tarif et Bon de Commande GMS - VRAC - MAJ Aout 2026.xlsx",
        "dst": "Tarif et Bon de Commande GMS NOEL 2026 - VRAC.xlsx",
        "sheet": "GMS NOEL 2026 VRAC",
        "titre": "TARIF & BON DE COMMANDE — GMS NOËL 2026 VRAC",
        "last_data": 32,
    },
]
SOUS_TITRE = "Offre de Noël 2026 — tarif août 2026"
NCOL = 15  # A..O


def style_of(ws, row):
    return [copy(ws.cell(row=row, column=c)._style) for c in range(1, NCOL + 1)]


def apply_style(ws, row, styles):
    for c in range(1, NCOL + 1):
        ws.cell(row=row, column=c)._style = copy(styles[c - 1])


def clear(ws, row):
    for c in range(1, NCOL + 1):
        ws.cell(row=row, column=c).value = None


for job in JOBS:
    src = os.path.join(BASE, job["src"])
    wb = openpyxl.load_workbook(src)
    ws = wb.active

    last = job["last_data"]
    old_sub = last + 1          # ligne « Sous-total »
    old_tot = old_sub + 2       # 1re ligne du bloc totaux (Total HTVA)

    # styles à réutiliser avant de toucher quoi que ce soit
    st_data = style_of(ws, last)
    st_head = style_of(ws, 11)
    st_sub = style_of(ws, old_sub)
    st_tot = [style_of(ws, old_tot + i) for i in range(4)]
    cond_txt = ws.cell(row=old_tot, column=2).value
    tot_lbl = [ws.cell(row=old_tot + i, column=13).value for i in range(4)]

    # démonter l'ancien bloc bas (le merge des conditions doit suivre le décalage)
    for rng in [str(r) for r in ws.merged_cells.ranges]:
        if rng.startswith("B"):
            ws.unmerge_cells(rng)
    clear(ws, old_sub)
    for i in range(4):
        clear(ws, old_tot + i)

    # nouvelles positions : 2 lignes insérées (titre de section + produit)
    r_head = old_sub
    r_prod = r_head + 1
    r_sub = r_prod + 1
    r_tot = r_sub + 2

    # 1. bandeau de section
    apply_style(ws, r_head, st_head)
    ws.cell(row=r_head, column=1).value = TITRE_SECTION
    ws.merge_cells(start_row=r_head, start_column=1, end_row=r_head, end_column=NCOL)

    # 2. la ligne produit
    apply_style(ws, r_prod, st_data)
    for col, val in NOEL.items():
        ws[f"{col}{r_prod}"] = val
    ws.cell(row=r_prod, column=14).value = f"=M{r_prod}*I{r_prod}"

    # 3. sous-total, recalé sur la nouvelle dernière ligne
    apply_style(ws, r_sub, st_sub)
    ws.cell(row=r_sub, column=1).value = "Sous-total"
    ws.cell(row=r_sub, column=13).value = f"=SUM(M12:M{r_prod})"
    ws.cell(row=r_sub, column=14).value = f"=SUM(N12:N{r_prod})"

    # 4. bloc totaux + conditions
    for i in range(4):
        apply_style(ws, r_tot + i, st_tot[i])
        ws.cell(row=r_tot + i, column=13).value = tot_lbl[i]
    ws.cell(row=r_tot, column=2).value = cond_txt
    ws.merge_cells(start_row=r_tot, start_column=2, end_row=r_tot + 3, end_column=11)
    ws.cell(row=r_tot, column=14).value = f"=N{r_sub}"
    ws.cell(row=r_tot + 1, column=14).value = f"=IF(N{r_tot}>=240,0,10)"
    ws.cell(row=r_tot + 2, column=14).value = f"=N{r_tot}*0.06"
    ws.cell(row=r_tot + 3, column=14).value = f"=SUM(N{r_tot}:N{r_tot + 2})"

    # 5. en-tête du document
    ws["A8"] = job["titre"]
    ws["A9"] = SOUS_TITRE
    ws.title = job["sheet"]

    dst = os.path.join(BASE, job["dst"])
    wb.save(dst)
    print(f"OK  {job['dst']}  (feuille « {job['sheet'] }», produit en ligne {r_prod}, "
          f"sous-total {r_sub}, totaux {r_tot}-{r_tot + 3})")

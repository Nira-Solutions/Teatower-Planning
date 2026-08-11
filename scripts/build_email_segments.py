"""
build_email_segments.py — Extraction Odoo des e-mails clients par segment commercial.

Produit un classeur Excel avec 3 segments EXCLUSIFS (un e-mail n'apparait que dans
un seul onglet, pour ne jamais envoyer 2 fois le meme message a la meme boite) :

    GMS        = tags Odoo  GMS (27), Canal GMS (88)
    REVENDEUR  = tags Odoo  Revendeur (28), Grossiste (32), IG (34),
                            MC-202606-REVENDEUR-DORMANT (91), rev (81)
    HORECA     = tags Odoo  HoReCA (26), Horeca Vrac & infu (31),
                            Horeca VIA Grossiste (33), Horeca Lost (79),
                            Canal Horeca (84), MC-202606-HORECA-DORMANT (90)

Priorite en cas de multi-tag : GMS > REVENDEUR > HORECA.
Un 4e onglet "B2B Entreprises" isole les comptes tagges uniquement
"Canal B2B Direct" (85) / Institution (29, 30) : entreprises et institutions qui
achetent pour leur consommation propre, ce ne sont PAS des revendeurs.

Le tag "Magasins" (80) est VOLONTAIREMENT ignore : c'est de la clientele B2C
boutique, pas un canal pro.

GRANULARITE — le point delicat. Chez Delhaize Le Lion S.A et Carrefour Belgium,
les magasins affilies sont des ADRESSES ENFANTS de la maison mere : rouler jusqu'a
`commercial_partner_id` ecraserait 71 Delhaize + 31 Carrefour en 2 lignes. Le
script detecte donc automatiquement les tetes de groupe (>= 8 fiches taggees sous
la meme societe) et redescend d'un cran : l'unite devient le magasin, pas le
groupe. Partout ailleurs, l'unite reste la societe et ses contacts sont regroupes
dessous.

Lecture seule cote Odoo.
Sortie : <OUT_DIR>/Emails_segments_B2B_<YYYY-MM-DD>.xlsx

Usage : python build_email_segments.py [--out DOSSIER]
"""

import argparse
import os
import re
import xmlrpc.client
from collections import Counter, defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

URL = "https://tea-tree.odoo.com"
DB = "tsc-be-tea-tree-main-18515272"
USER = "nicolas.raes@teatower.com"
PASSWORD = os.environ.get("ODOO_PWD", "Teatower123")

# --- Segmentation ---------------------------------------------------------
TAG_NAMES = {
    26: "HoReCA", 27: "GMS", 28: "Revendeur", 29: "Institution",
    30: "Institution TT", 31: "Horeca Vrac & infu", 32: "Grossiste",
    33: "Horeca VIA Grossiste", 34: "IG", 79: "Horeca Lost", 80: "Magasins",
    81: "rev", 82: "Faire", 83: "B2B", 84: "Canal Horeca",
    85: "Canal B2B Direct", 86: "Canal DTC Shopify", 87: "Canal DTC Amazon",
    88: "Canal GMS", 90: "HORECA-DORMANT", 91: "REVENDEUR-DORMANT",
}
SEG_TAGS = [
    ("GMS", [27, 88]),
    ("REVENDEUR", [28, 32, 34, 81, 91]),
    ("HORECA", [26, 31, 33, 79, 84, 90]),
    ("B2B_ENTREPRISE", [29, 30, 85]),
]
SEG_ORDER = ["HORECA", "REVENDEUR", "GMS", "B2B_ENTREPRISE"]
SEG_LABEL = {
    "HORECA": "Horeca",
    "REVENDEUR": "Revendeurs B2B",
    "GMS": "GMS",
    "B2B_ENTREPRISE": "B2B Entreprises",
}
ALL_SEG_TAGS = sorted({t for _, ts in SEG_TAGS for t in ts})

# Une societe portant au moins ce nombre de fiches taggees est une tete de
# groupe : on descend d'un cran pour garder la granularite magasin.
GROUP_THRESHOLD = 8

# Equipes de vente exclues du calcul de CA B2B (Shopify, caisse, Amazon).
NON_B2B_TEAMS = [3, 4, 5, 6, 7, 8]

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-']+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
INTERNAL_DOMAINS = ("teatower.com", "tea-tree.odoo.com", "nirasolutions")
JUNK_MARKERS = ("noreply", "no-reply", "example.com", "marketplace.amazon")

P_FIELDS = ["id", "name", "email", "phone", "mobile", "function", "city",
            "zip", "country_id", "category_id", "is_company", "parent_id",
            "commercial_partner_id", "active", "comment", "type"]


def connect():
    uid = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/common").authenticate(
        DB, USER, PASSWORD, {})
    if not uid:
        raise SystemExit("Authentification Odoo refusee")
    models = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/object")

    def call(model, method, *args, **kwargs):
        return models.execute_kw(DB, uid, PASSWORD, model, method,
                                 list(args), kwargs)

    return call


def clean_emails(raw):
    """Un champ Odoo peut contenir plusieurs adresses ('a@b.be; c@d.be'),
    ou carrement un numero de telephone. On ne garde que du valide."""
    out = []
    for hit in EMAIL_RE.findall(raw or ""):
        e = hit.strip().strip(".,;").lower()
        if any(d in e for d in INTERNAL_DOMAINS):
            continue
        if any(j in e for j in JUNK_MARKERS):
            continue
        if e not in out:
            out.append(e)
    return out


def segment_of(tag_ids):
    """Segment exclusif : GMS > REVENDEUR > HORECA > B2B_ENTREPRISE."""
    tags = set(tag_ids)
    for seg, seg_tags in SEG_TAGS:
        if tags & set(seg_tags):
            return seg
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(Path.home() / "OneDrive" /
                                        "Teatower-Direction" / "direction"))
    args = ap.parse_args()

    call = connect()
    today = date.today()
    d12 = (today - timedelta(days=365)).isoformat()
    d24 = (today - timedelta(days=730)).isoformat()

    def fetch(ids):
        """Lit les partenaires manquants, archives compris."""
        todo = [i for i in ids if i and i not in M]
        for k in range(0, len(todo), 400):
            for p in call("res.partner", "search_read",
                          [["id", "in", todo[k:k + 400]]], fields=P_FIELDS,
                          limit=1000, context={"active_test": False}):
                M[p["id"]] = p

    M = {}
    print("1/6  Lecture des fiches taggees...")
    tagged = call("res.partner", "search_read",
                  [["category_id", "in", ALL_SEG_TAGS]],
                  fields=P_FIELDS, limit=8000)
    for p in tagged:
        M[p["id"]] = p
    print(f"     {len(tagged)} fiches")

    print("2/6  Remontee des maisons meres + descente sur les contacts...")
    fetch([(p["parent_id"] or [0])[0] for p in tagged] +
          [(p["commercial_partner_id"] or [0])[0] for p in tagged])
    comm_ids = sorted({(p["commercial_partner_id"] or [p["id"]])[0]
                       for p in tagged})
    for k in range(0, len(comm_ids), 150):
        for p in call("res.partner", "search_read",
                      [["id", "child_of", comm_ids[k:k + 150]]],
                      fields=P_FIELDS, limit=20000,
                      context={"active_test": False}):
            M.setdefault(p["id"], p)
    print(f"     {len(comm_ids)} societes racines, {len(M)} fiches en memoire")

    # --- Tetes de groupe (Delhaize Le Lion, Carrefour Belgium...) ----------
    per_comm = Counter((p["commercial_partner_id"] or [p["id"]])[0]
                       for p in tagged)
    groups = {cid for cid, n in per_comm.items() if n >= GROUP_THRESHOLD}
    print("3/6  Tetes de groupe detectees : " +
          (", ".join(f"{M[g]['name']} ({per_comm[g]})" for g in groups)
           or "aucune"))

    def unit_of(pid):
        """Unite commerciale d'une fiche : la societe, sauf sous une tete de
        groupe ou l'on redescend au magasin (enfant direct du groupe)."""
        p = M.get(pid)
        if not p:
            return pid
        cid = (p["commercial_partner_id"] or [pid])[0]
        if cid not in groups:
            return cid
        cur = pid
        while True:
            q = M.get(cur)
            par = (q or {}).get("parent_id")
            if not q or not par:
                return cur
            if par[0] == cid:
                return cur
            cur = par[0]

    unit_tags = defaultdict(set)
    for p in tagged:
        unit_tags[unit_of(p["id"])].update(p["category_id"])
    print(f"     {len(unit_tags)} unites commerciales")

    print("4/6  Historique commercial (SO confirmees, hors Shopify/caisse)...")
    hist = defaultdict(lambda: {"last": None, "n24": 0, "ca12": 0.0})
    for k in range(0, len(comm_ids), 150):
        for o in call("sale.order", "search_read",
                      [["state", "in", ["sale", "done"]],
                       ["team_id", "not in", NON_B2B_TEAMS],
                       ["partner_id", "child_of", comm_ids[k:k + 150]]],
                      fields=["partner_id", "date_order", "amount_untaxed"],
                      limit=40000):
            h = hist[unit_of(o["partner_id"][0])]
            d = o["date_order"][:10]
            if h["last"] is None or d > h["last"]:
                h["last"] = d
            if d >= d24:
                h["n24"] += 1
            if d >= d12:
                h["ca12"] += o["amount_untaxed"]

    print("5/6  Construction des segments...")
    emails_of_unit = defaultdict(list)
    for p in M.values():
        for e in clean_emails(p.get("email")):
            u = unit_of(p["id"])
            if u in unit_tags:
                emails_of_unit[u].append((e, p))

    rows = defaultdict(list)
    seen = set()          # exclusivite globale des adresses
    units = sorted(unit_tags, key=lambda u: -hist[u]["ca12"])
    for u in units:
        seg = segment_of(unit_tags[u])
        if not seg:
            continue
        unit = M.get(u)
        if not unit:
            continue
        comm = (unit.get("commercial_partner_id") or [u, ""])
        groupe = comm[1] if comm[0] in groups and comm[0] != u else ""
        comment = (unit.get("comment") or "").upper()
        h = hist[u]
        if not unit.get("active", True):
            statut = "Archive"
        elif "ARRET" in comment or "ARRÊT" in comment or "NO-MERCH" in comment:
            statut = "ARRET / no-merch"
        elif h["last"] and h["last"] >= d12:
            statut = "Actif 12m"
        elif h["last"] and h["last"] >= d24:
            statut = "Dormant 12-24m"
        elif h["last"]:
            statut = "Inactif >24m"
        else:
            statut = "Jamais commande"
        tags_lbl = ", ".join(sorted(TAG_NAMES.get(t, str(t))
                                    for t in unit_tags[u]))
        nom = unit["name"] or (unit.get("parent_id") or ["", "(sans nom)"])[1]

        for e, p in sorted(emails_of_unit.get(u, []),
                           key=lambda x: (x[1]["id"] != u, x[0])):
            if e in seen:
                continue
            seen.add(e)
            contact = "" if p["id"] == u else (p["name"] or "")
            rows[seg].append({
                "email": e,
                "client": nom,
                "groupe": groupe,
                "contact": contact,
                "fonction": p.get("function") or "",
                "tel": p.get("phone") or p.get("mobile") or "",
                "cp": p.get("zip") or unit.get("zip") or "",
                "ville": p.get("city") or unit.get("city") or "",
                "pays": ((p.get("country_id") or unit.get("country_id"))
                         or ["", ""])[1],
                "statut": statut,
                "last": h["last"] or "",
                "n24": h["n24"],
                "ca12": round(h["ca12"], 2),
                "tags": tags_lbl,
                "pid": u,
            })

    for seg in SEG_ORDER:
        rows[seg].sort(key=lambda r: (r["statut"] != "Actif 12m",
                                      -r["ca12"], r["client"].lower()))
        print(f"     {SEG_LABEL[seg]:18s} : {len(rows[seg]):4d} e-mails "
              f"/ {len({r['pid'] for r in rows[seg]}):4d} clients")

    print("6/6  Ecriture du classeur...")
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"Emails_segments_B2B_{today.isoformat()}.xlsx"
    write_workbook(path, rows, today)
    print(f"OK -> {path}")


# --------------------------------------------------------------------------
GREEN = "0B6E4F"
HEAD_FILL = PatternFill("solid", fgColor=GREEN)
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
COLS = [("E-mail", 36), ("Client", 34), ("Groupe", 20), ("Contact", 22),
        ("Fonction", 16), ("Telephone", 16), ("CP", 8), ("Ville", 18),
        ("Pays", 12), ("Statut", 16), ("Derniere commande", 17),
        ("Cdes 24m", 10), ("CA 12m HT", 12), ("Tags Odoo", 38), ("ID", 8)]


def write_workbook(path, rows, today):
    wb = Workbook()
    wb.remove(wb.active)

    # --- Onglet COPIER : blocs prets a coller -----------------------------
    ws = wb.create_sheet("COPIER")
    ws.sheet_properties.tabColor = GREEN
    ws["A1"] = f"E-mails clients par segment — extrait Odoo du {today:%d/%m/%Y}"
    ws["A1"].font = Font(bold=True, size=14, color=GREEN)
    ws["A2"] = ("Clique une cellule, copie (Ctrl+C), colle dans le champ "
                "destinataires. Separateur : point-virgule.")
    ws["A2"].font = Font(italic=True, size=10, color="666666")
    ws["A3"] = ("Segments exclusifs : une adresse n'apparait que dans un seul "
                "onglet. Colonne D = clients ayant commande dans les 12 mois.")
    ws["A3"].font = Font(italic=True, size=10, color="666666")
    for col, lbl in ((1, "Segment"), (2, "Nb e-mails"), (3, "Tous les e-mails"),
                     (4, "Actifs 12 mois"), (5, "Nb actifs")):
        c = ws.cell(row=5, column=col, value=lbl)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(horizontal="center")
    r = 6
    for seg in SEG_ORDER:
        allm = [x["email"] for x in rows[seg]]
        act = [x["email"] for x in rows[seg] if x["statut"] == "Actif 12m"]
        ws.cell(row=r, column=1, value=SEG_LABEL[seg]).font = Font(bold=True)
        ws.cell(row=r, column=2, value=len(allm)).alignment = Alignment(
            horizontal="center")
        ws.cell(row=r, column=3, value="; ".join(allm))
        ws.cell(row=r, column=4, value="; ".join(act))
        ws.cell(row=r, column=5, value=len(act)).alignment = Alignment(
            horizontal="center")
        ws.row_dimensions[r].height = 26
        r += 1
    for col, w in ((1, 20), (2, 12), (3, 60), (4, 60), (5, 12)):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A6"

    # --- Un onglet par segment --------------------------------------------
    for seg in SEG_ORDER:
        ws = wb.create_sheet(SEG_LABEL[seg])
        for i, (lbl, w) in enumerate(COLS, start=1):
            c = ws.cell(row=1, column=i, value=lbl)
            c.fill, c.font = HEAD_FILL, HEAD_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22
        for j, x in enumerate(rows[seg], start=2):
            vals = [x["email"], x["client"], x["groupe"], x["contact"],
                    x["fonction"], x["tel"], x["cp"], x["ville"], x["pays"],
                    x["statut"], x["last"], x["n24"], x["ca12"], x["tags"],
                    x["pid"]]
            red = x["statut"] in ("ARRET / no-merch", "Archive")
            for i, v in enumerate(vals, start=1):
                c = ws.cell(row=j, column=i, value=v)
                if i == 13:
                    c.number_format = '#,##0 "€"'
                if i in (7, 12, 15):
                    c.alignment = Alignment(horizontal="center")
                if red:
                    c.font = Font(color="B00020")
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = \
            f"A1:{get_column_letter(len(COLS))}{len(rows[seg]) + 1}"

    wb.save(path)


if __name__ == "__main__":
    main()
